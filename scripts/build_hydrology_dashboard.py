# -*- coding: utf-8 -*-
"""Build the hydrology cost-analysis dashboard from K-water survey workbooks."""

from __future__ import annotations

import json
import re
import sqlite3
from collections import Counter, defaultdict
from datetime import date
from pathlib import Path
from statistics import mean
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "docs" / "과업수행 관련 자료_한국수자원조사기술원"
OUTPUT_HTML = ROOT / "docs" / "수문조사_원가분석_대시보드.html"
DATA_DIR = ROOT / "docs" / "data"
YEARS = ["2021년", "2022년", "2023년", "2024년", "2025년", "2026년"]
REGION_RANGES = [
    ("한강권역", 1, 97),
    ("낙동강권역", 98, 201),
    ("금강권역", 202, 278),
    ("영산강권역", 279, 355),
]


def clean(value: Any) -> Any:
    if isinstance(value, str):
        return value.replace("\n", " ").strip()
    return value


def number(value: Any, default: float = 0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else default


def won_from_text(value: Any) -> int:
    return int(round(number(value)))


def workbook(needle: str) -> Path:
    matches = [
        path
        for path in SOURCE_DIR.rglob("*.xlsx")
        if needle in path.name and not path.name.startswith("~$")
    ]
    if not matches:
        raise FileNotFoundError(f"Workbook not found: {needle}")
    return sorted(matches)[0]


def region_for_station(no: int) -> str:
    for name, start, end in REGION_RANGES:
        if start <= no <= end:
            return name
    return "미분류"


def expand_range(value: Any) -> list[int]:
    if isinstance(value, (int, float)):
        return [int(value)]
    nums = [int(v) for v in re.findall(r"\d+", str(value or ""))]
    if len(nums) >= 2:
        return list(range(nums[0], nums[1] + 1))
    if len(nums) == 1:
        return [nums[0]]
    return []


def rows_by_year(ws: Any, row_start: int, row_end: int, label_col: int, first_year_col: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_no in range(row_start, row_end + 1):
        label = clean(ws.cell(row=row_no, column=label_col).value)
        if not label:
            continue
        rows.append(
            {
                "item": label,
                "values": {
                    year[:4]: int(round(number(ws.cell(row=row_no, column=first_year_col + idx).value)))
                    for idx, year in enumerate(YEARS)
                },
            }
        )
    return rows


def parse_budget_and_units() -> dict[str, Any]:
    path = workbook("예산 및 단가 현황_v2")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    counts = rows_by_year(wb["현황"], 5, 11, 3, 4)
    counts_after = rows_by_year(wb["현황"], 5, 11, 12, 13)
    unit_prices = rows_by_year(wb["단가"], 5, 11, 3, 4)
    business_budget = rows_by_year(wb["항목별예산"], 14, 19, 2, 3)
    business_budget_with_carryover = rows_by_year(wb["항목별예산"], 14, 19, 10, 11)

    flow_rows: list[dict[str, Any]] = []
    ws = wb["유량"]
    for row_no in range(6, 12):
        flow_rows.append(
            {
                "year": str(ws.cell(row=row_no, column=2).value)[:4],
                "sites": int(number(ws.cell(row=row_no, column=3).value)),
                "unit_price_million": int(number(ws.cell(row=row_no, column=4).value)),
                "total_million": int(number(ws.cell(row=row_no, column=5).value)),
                "org_million": int(number(ws.cell(row=row_no, column=6).value)),
                "org_pct": number(ws.cell(row=row_no, column=7).value),
                "business_million": int(number(ws.cell(row=row_no, column=8).value)),
                "business_pct": number(ws.cell(row=row_no, column=9).value),
            }
        )

    total_budget: list[dict[str, Any]] = []
    ws = wb["총예산_02"]
    for row_no in range(6, 12):
        total_budget.append(
            {
                "year": str(ws.cell(row=row_no, column=2).value)[:4],
                "total_million": int(number(ws.cell(row=row_no, column=3).value)),
                "org_million": int(number(ws.cell(row=row_no, column=4).value)),
                "org_pct": number(ws.cell(row=row_no, column=5).value),
                "business_million": int(number(ws.cell(row=row_no, column=6).value)),
                "business_pct": number(ws.cell(row=row_no, column=7).value),
            }
        )

    auto_rows: list[dict[str, Any]] = []
    ws = wb["자동유량"]
    current_year = "2021"
    for row_no in range(7, 30):
        year_cell = ws.cell(row=row_no, column=2).value
        if year_cell:
            current_year = str(year_cell)[:4]
        item = clean(ws.cell(row=row_no, column=3).value)
        if item in {"운영", "설치", "유지관리"}:
            auto_rows.append(
                {
                    "year": current_year,
                    "item": item,
                    "quantity": int(number(ws.cell(row=row_no, column=4).value)),
                    "unit_price_million": int(number(ws.cell(row=row_no, column=5).value)),
                    "total_million": int(number(ws.cell(row=row_no, column=6).value)),
                    "business_million": int(number(ws.cell(row=row_no, column=9).value)),
                }
            )

    return {
        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
        "counts": counts,
        "counts_after": counts_after,
        "unit_prices": unit_prices,
        "business_budget": business_budget,
        "business_budget_with_carryover": business_budget_with_carryover,
        "flow_rows": flow_rows,
        "total_budget": total_budget,
        "auto_rows": auto_rows,
    }


def parse_flow_business_breakdown() -> dict[str, Any]:
    path = workbook("세부예산 현황_v1")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["유량"]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if clean(row[2]) != "소계" or clean(row[1]) == "합계":
            continue
        value = int(number(row[13]))
        if value:
            rows.append({"item": clean(row[1]), "amount_won": value})
    total = sum(row["amount_won"] for row in rows)
    return {
        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
        "year": "2026",
        "rows": rows,
        "total_won": total,
    }


def parse_stations_and_traffic() -> dict[str, Any]:
    station_path = workbook("조사지점 현황")
    traffic_path = workbook("교통비")

    wb = openpyxl.load_workbook(station_path, data_only=True, read_only=True)
    ws = wb.active
    stations: dict[int, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        no = row[1]
        if isinstance(no, int):
            stations[no] = {
                "no": no,
                "name": clean(row[2]),
                "address": clean(row[3]),
                "region": region_for_station(no),
            }

    wb = openpyxl.load_workbook(traffic_path, data_only=True, read_only=True)
    ws = wb.active
    traffic_ranges: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        station_numbers = expand_range(row[0])
        fare = won_from_text(row[4])
        if not station_numbers or not fare:
            continue
        rec = {
            "range": clean(row[0]),
            "place": clean(row[1]),
            "mode": clean(row[2]),
            "hub": clean(row[3]),
            "fare": fare,
            "start": min(station_numbers),
            "end": max(station_numbers),
            "region": region_for_station(min(station_numbers)),
        }
        traffic_ranges.append(rec)
        for no in station_numbers:
            if no in stations:
                stations[no].update(
                    {
                        "fare": fare,
                        "mode": rec["mode"],
                        "hub": rec["hub"],
                        "traffic_place": rec["place"],
                    }
                )

    return {
        "station_source": str(station_path.relative_to(ROOT)).replace("\\", "/"),
        "traffic_source": str(traffic_path.relative_to(ROOT)).replace("\\", "/"),
        "stations": [stations[key] for key in sorted(stations)],
        "traffic_ranges": traffic_ranges,
    }


def parse_equipment() -> dict[str, Any]:
    path = workbook("장비 관련")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["보유장비 관련"]
    rows: list[dict[str, Any]] = []
    current_group = ""
    for row in ws.iter_rows(min_row=5, values_only=True):
        group = clean(row[1])
        sub = clean(row[2])
        variant = clean(row[3])
        if group:
            current_group = group
        if current_group == "합계":
            continue
        owned = int(number(row[4], 0))
        if not current_group or not owned:
            continue
        name = " ".join(str(part) for part in [current_group, sub, variant] if part)
        rows.append(
            {
                "name": name,
                "group": current_group,
                "owned": owned,
                "standard": clean(row[5]) or "-",
                "life_elapsed_2026": int(number(row[6], 0)),
                "life_elapsed_2027": int(number(row[7], 0)),
                "purchase_plan_2027": int(number(row[8], 0)),
                "unit_price_won": int(number(row[9], 0)) if row[9] not in (None, "-") else None,
            }
        )

    calibration = []
    ws = wb["월 검·교정 비용"]
    for col in range(3, 15):
        month = clean(ws.cell(row=3, column=col).value)
        amount_thousand = number(ws.cell(row=4, column=col).value, 0)
        calibration.append({"month": month, "amount_won": int(round(amount_thousand * 1000))})

    def avg_price(group: str) -> float:
        values = [
            row["unit_price_won"]
            for row in rows
            if group in row["group"] and row["unit_price_won"]
        ]
        return mean(values) if values else 0

    core_kit = round(
        avg_price("ADCP")
        + avg_price("무선조종 보트")
        + avg_price("전자파 표면유속계")
        + avg_price("측량장비")
    )

    return {
        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
        "rows": rows,
        "total_owned": sum(row["owned"] for row in rows),
        "calibration": calibration,
        "calibration_annual_won": sum(row["amount_won"] for row in calibration),
        "core_kit_price_won": core_kit,
    }


def parse_vehicle_ops() -> dict[str, Any]:
    path = workbook("업무차량")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    ws = wb["업무차량"]
    vehicles: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if clean(row[1]) and isinstance(row[3], (int, float)):
            vehicles.append(
                {
                    "type": clean(row[1]),
                    "eco_type": clean(row[2]),
                    "monthly_rent_won": int(number(row[3])),
                }
            )
    counts = Counter(row["type"] for row in vehicles)

    ws = wb["전기&주유비"]
    energy_rows: list[dict[str, Any]] = []
    months = [clean(ws.cell(row=3, column=col).value) for col in range(3, 9)]
    for row_no in [4, 5]:
        energy_rows.append(
            {
                "item": clean(ws.cell(row=row_no, column=2).value),
                "values": {
                    month: int(number(ws.cell(row=row_no, column=idx + 3).value))
                    for idx, month in enumerate(months)
                },
            }
        )

    return {
        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
        "vehicle_count": len(vehicles),
        "avg_monthly_rent_won": round(mean(row["monthly_rent_won"] for row in vehicles)),
        "total_monthly_rent_won": sum(row["monthly_rent_won"] for row in vehicles),
        "type_counts": [{"type": key, "count": value} for key, value in counts.most_common()],
        "energy_rows": energy_rows,
    }


def parse_rent_ops() -> dict[str, Any]:
    path = workbook("임대 관련")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    def annual_by_region(sheet_name: str, start_row: int, end_row: int) -> dict[str, int]:
        ws = wb[sheet_name]
        result: dict[str, int] = {}
        for row in ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True):
            label = clean(row[1])
            if not label:
                continue
            result[label.replace("실", "")] = int(round(number(row[14]) * 1000))
        return result

    container = annual_by_region("컨테이너(창고)_2025년 기준", 5, 8)
    parking = annual_by_region("주차시설_2025년 기준", 6, 9)
    survey_equipment_rent = int(
        round(number(wb["측량장비 대여_2025년 기준"].cell(row=4, column=14).value) * 1000)
    )
    return {
        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
        "container_annual_won": container,
        "parking_annual_won": parking,
        "survey_equipment_rent_won": survey_equipment_rent,
    }


def parse_budget_summary_records() -> list[dict[str, Any]]:
    path = workbook("예산 및 단가 현황_v2")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["항목별예산"]
    records: list[dict[str, Any]] = []
    for row_no in range(5, 11):
        item = clean(ws.cell(row=row_no, column=2).value)
        item_with_carryover = clean(ws.cell(row=row_no, column=10).value) or item
        if not item:
            continue
        for idx, year in enumerate(YEARS):
            execution = int(number(ws.cell(row=row_no, column=3 + idx).value))
            with_carryover = int(number(ws.cell(row=row_no, column=11 + idx).value))
            records.append(
                {
                    "item": item,
                    "item_with_carryover": item_with_carryover,
                    "year": year[:4],
                    "execution_budget_won": execution,
                    "balance_reflected_budget_won": with_carryover,
                    "balance_delta_won": with_carryover - execution,
                    "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                    "source_sheet": "항목별예산",
                    "source_row": row_no,
                }
            )
    return records


def parse_budget_detail_records() -> list[dict[str, Any]]:
    path = workbook("세부예산 현황_v1")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    records: list[dict[str, Any]] = []
    for sheet_name in ["유량", "유사량", "토양수분량", "증발산량", "자동유량"]:
        ws = wb[sheet_name]
        current_group = ""
        for row_no, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
            group = clean(row[1])
            stat_item = clean(row[2])
            if group:
                current_group = str(group)
            if current_group == "합계" and not stat_item:
                stat_item = "합계"
            if not current_group or not stat_item:
                continue
            for idx, year in enumerate(YEARS):
                execution = int(number(row[3 + idx * 2]))
                balance = int(number(row[4 + idx * 2]))
                if not execution and not balance:
                    continue
                records.append(
                    {
                        "measurement_item": sheet_name,
                        "budget_group": current_group,
                        "stat_item": stat_item,
                        "year": year[:4],
                        "execution_budget_won": execution,
                        "carryover_balance_budget_won": balance,
                        "budget_with_balance_won": execution + balance,
                        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                        "source_sheet": sheet_name,
                        "source_row": row_no,
                    }
                )
    return records


def parse_residual_transactions() -> list[dict[str, Any]]:
    path = workbook("세부예산 현황_v1")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["잔액"]
    records: list[dict[str, Any]] = []
    project_name = ""
    for row_no, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        if clean(row[1]):
            project_name = clean(row[1])
        approval_date = row[3]
        budget_group = clean(row[7])
        amount = int(number(row[9]))
        if not approval_date or not budget_group or budget_group == "편성목":
            continue
        records.append(
            {
                "project_name": project_name,
                "approval_date": approval_date.isoformat() if hasattr(approval_date, "isoformat") else str(approval_date),
                "resolution_title": clean(row[4]),
                "budget_department": clean(row[5]),
                "resolution_department": clean(row[6]),
                "budget_group": budget_group,
                "stat_item": clean(row[8]),
                "amount_won": amount,
                "classification_item": clean(row[10]),
                "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                "source_sheet": "잔액",
                "source_row": row_no,
            }
        )
    return records


def parse_unit_price_records() -> list[dict[str, Any]]:
    path = workbook("예산 및 단가 현황_v2")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    status_ws = wb["현황"]
    price_ws = wb["단가"]
    records: list[dict[str, Any]] = []
    current_group_before = ""
    current_group_after = ""
    current_group_price = ""
    for row_no in range(5, 12):
        group_before = clean(status_ws.cell(row=row_no, column=2).value)
        group_after = clean(status_ws.cell(row=row_no, column=11).value)
        group_price = clean(price_ws.cell(row=row_no, column=2).value)
        if group_before:
            current_group_before = group_before
        if group_after:
            current_group_after = group_after
        if group_price:
            current_group_price = group_price
        item_before = clean(status_ws.cell(row=row_no, column=3).value)
        item_after = clean(status_ws.cell(row=row_no, column=12).value)
        item_price = clean(price_ws.cell(row=row_no, column=3).value)
        for idx, year in enumerate(YEARS):
            unit_price = int(number(price_ws.cell(row=row_no, column=4 + idx).value))
            if item_before:
                count_before = int(number(status_ws.cell(row=row_no, column=4 + idx).value))
                records.append(
                    {
                        "version": "before",
                        "group": current_group_before,
                        "item": item_before,
                        "year": year[:4],
                        "count": count_before,
                        "unit_price_million_won": unit_price,
                        "total_million_won": count_before * unit_price,
                        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                        "source_sheet": "현황/단가",
                        "source_row": row_no,
                    }
                )
            if item_after:
                count_after = int(number(status_ws.cell(row=row_no, column=13 + idx).value))
                records.append(
                    {
                        "version": "after",
                        "group": current_group_after or current_group_price,
                        "item": item_after,
                        "year": year[:4],
                        "count": count_after,
                        "unit_price_million_won": unit_price,
                        "total_million_won": count_after * unit_price,
                        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                        "source_sheet": "현황/단가",
                        "source_row": row_no,
                    }
                )
    return records


def parse_total_budget_breakdown_records() -> list[dict[str, Any]]:
    path = workbook("예산 및 단가 현황_v2")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["총예산_01"]
    item_labels = ["유량", "유사량", "토양수분량", "증발산량", "운영", "설치", "유지관리"]
    records: list[dict[str, Any]] = []
    current_year_before = ""
    current_year_after = ""
    for row_no in range(6, 24):
        year_before = clean(ws.cell(row=row_no, column=2).value)
        category_before = clean(ws.cell(row=row_no, column=3).value)
        if year_before:
            current_year_before = str(year_before)[:4]
        if current_year_before and category_before:
            for idx, item in enumerate(["계"] + item_labels):
                value = int(number(ws.cell(row=row_no, column=4 + idx).value))
                records.append(
                    {
                        "year": current_year_before,
                        "row_category": category_before,
                        "item": item,
                        "amount_million_won": value,
                        "carryover_included": False,
                        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                        "source_sheet": "총예산_01",
                        "source_row": row_no,
                    }
                )

        year_after = clean(ws.cell(row=row_no, column=13).value)
        category_after = clean(ws.cell(row=row_no, column=14).value)
        if year_after:
            current_year_after = str(year_after)[:4]
        if current_year_after and category_after:
            for idx, item in enumerate(["계"] + item_labels):
                value = int(number(ws.cell(row=row_no, column=15 + idx).value))
                records.append(
                    {
                        "year": current_year_after,
                        "row_category": category_after,
                        "item": item,
                        "amount_million_won": value,
                        "carryover_included": True,
                        "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                        "source_sheet": "총예산_01",
                        "source_row": row_no,
                    }
                )
    return records


def build_cost_category_mapping() -> list[dict[str, str]]:
    return [
        {"standard_cost_category": "인건비", "budget_group": "인건비", "stat_item": "*", "survey_type": "기본 현장 유량측정", "allocation_rule": "직접 연결"},
        {"standard_cost_category": "장비비", "budget_group": "유형자산", "stat_item": "자산취득비", "survey_type": "기본 현장 유량측정", "allocation_rule": "장비 내용연수·투입횟수 배분"},
        {"standard_cost_category": "장비비", "budget_group": "운영비", "stat_item": "시설장비유지비", "survey_type": "자동유량 관측소 운영", "allocation_rule": "운영 점검비 배분"},
        {"standard_cost_category": "여비", "budget_group": "여비", "stat_item": "국내여비", "survey_type": "기본 현장 유량측정", "allocation_rule": "조사지점 교통비 기반"},
        {"standard_cost_category": "보고서작성비", "budget_group": "연구용역비", "stat_item": "*", "survey_type": "공통", "allocation_rule": "자료분석·성과품 작성 역할 비용"},
        {"standard_cost_category": "간접비", "budget_group": "일반관리비", "stat_item": "*", "survey_type": "공통", "allocation_rule": "직접비 대비율"},
        {"standard_cost_category": "설치비", "budget_group": "건설비", "stat_item": "*", "survey_type": "자동유량 관측소 신규 설치", "allocation_rule": "설치 단가와 사업비 역산"},
    ]


def build_survey_cost_model(
    budget: dict[str, Any],
    flow_breakdown: dict[str, Any],
    equipment: dict[str, Any],
    vehicles: dict[str, Any],
    regions: list[dict[str, Any]],
) -> dict[str, Any]:
    role_rates = {
        "책임기술자": 56250,
        "현장조사원": 37500,
        "보조조사원": 27500,
        "자료분석원": 40000,
        "보고서작성자": 35000,
    }
    role_templates = {
        "basic_flow": [
            ("책임기술자", 1, 2.0, "인건비"),
            ("현장조사원", 2, 6.0, "인건비"),
            ("보조조사원", 1, 4.0, "인건비"),
            ("자료분석원", 1, 3.0, "인건비"),
            ("보고서작성자", 1, 2.0, "보고서작성비"),
        ],
        "auto_operation": [
            ("책임기술자", 1, 1.5, "인건비"),
            ("현장조사원", 1, 4.0, "인건비"),
            ("자료분석원", 1, 2.5, "인건비"),
            ("보고서작성자", 1, 1.5, "보고서작성비"),
        ],
        "auto_install": [
            ("책임기술자", 1, 6.0, "인건비"),
            ("현장조사원", 2, 8.0, "인건비"),
            ("보조조사원", 2, 8.0, "인건비"),
            ("자료분석원", 1, 4.0, "인건비"),
            ("보고서작성자", 1, 3.0, "보고서작성비"),
        ],
    }

    avg_round_trip_fare = round(mean(row["avg_fare"] for row in regions) * 2)
    vehicle_per_visit = round(vehicles["avg_monthly_rent_won"] / 22)
    equipment_per_visit = round(equipment["core_kit_price_won"] / 7 / 160)

    flow_2026 = next(row for row in budget["flow_rows"] if row["year"] == "2026")
    flow_frequency = 12
    basic_reverse_business = round(flow_breakdown["total_won"] / flow_2026["sites"] / flow_frequency)
    basic_reverse_total = round(flow_2026["total_million"] * 1_000_000 / flow_2026["sites"] / flow_frequency)

    def auto_row(item: str) -> dict[str, Any]:
        return next(row for row in budget["auto_rows"] if row["year"] == "2026" and row["item"] == item)

    auto_operating = auto_row("운영")
    auto_install = auto_row("설치")
    operation_frequency = 12
    operation_reverse_total = round(auto_operating["total_million"] * 1_000_000 / auto_operating["quantity"] / operation_frequency)
    operation_reverse_business = round(auto_operating["business_million"] * 1_000_000 / auto_operating["quantity"] / operation_frequency)
    install_reverse_total = round(auto_install["total_million"] * 1_000_000 / max(auto_install["quantity"], 1))
    install_reverse_business = round(auto_install["business_million"] * 1_000_000 / max(auto_install["quantity"], 1))

    unit_models = [
        {
            "survey_type_id": "basic_flow",
            "survey_type": "기본 현장 유량측정",
            "unit_scope": "1개 측정지점 1회 조사",
            "annual_frequency": flow_frequency,
            "site_count_basis": flow_2026["sites"],
            "reverse_total_cost_won": basic_reverse_total,
            "reverse_business_cost_won": basic_reverse_business,
            "equipment_cost_won": equipment_per_visit,
            "travel_cost_won": avg_round_trip_fare * 2,
            "vehicle_cost_won": vehicle_per_visit,
            "difficulty_level": "보통",
        },
        {
            "survey_type_id": "auto_operation",
            "survey_type": "자동유량 관측소 운영",
            "unit_scope": "1개 측정지점 1회 점검·자료검증",
            "annual_frequency": operation_frequency,
            "site_count_basis": auto_operating["quantity"],
            "reverse_total_cost_won": operation_reverse_total,
            "reverse_business_cost_won": operation_reverse_business,
            "equipment_cost_won": 250000,
            "travel_cost_won": avg_round_trip_fare,
            "vehicle_cost_won": vehicle_per_visit,
            "difficulty_level": "보통",
        },
        {
            "survey_type_id": "auto_install",
            "survey_type": "자동유량 관측소 신규 설치",
            "unit_scope": "1개 측정지점 1개소 설치",
            "annual_frequency": 1,
            "site_count_basis": auto_install["quantity"],
            "reverse_total_cost_won": install_reverse_total,
            "reverse_business_cost_won": install_reverse_business,
            "equipment_cost_won": round(install_reverse_total * 0.82),
            "travel_cost_won": avg_round_trip_fare * 4,
            "vehicle_cost_won": vehicle_per_visit * 3,
            "difficulty_level": "고난도",
        },
    ]

    roles: list[dict[str, Any]] = []
    category_rows: list[dict[str, Any]] = []
    for model in unit_models:
        survey_id = model["survey_type_id"]
        category_totals: defaultdict[str, int] = defaultdict(int)
        for role_name, headcount, hours, category in role_templates[survey_id]:
            hourly_rate = role_rates[role_name]
            labor_cost = round(headcount * hours * hourly_rate)
            roles.append(
                {
                    "survey_type_id": survey_id,
                    "survey_type": model["survey_type"],
                    "role_name": role_name,
                    "headcount": headcount,
                    "hours_per_person": hours,
                    "labor_grade": role_name,
                    "hourly_rate_won": hourly_rate,
                    "labor_cost_won": labor_cost,
                    "cost_category": category,
                    "basis_method": "역할별 직접 입력 + 예산 역산 비교",
                }
            )
            category_totals[category] += labor_cost
        category_totals["장비비"] += int(model["equipment_cost_won"])
        category_totals["여비"] += int(model["travel_cost_won"])
        category_totals["차량운영비"] += int(model["vehicle_cost_won"])
        direct = sum(category_totals.values())
        overhead = round(direct * 0.17)
        category_totals["간접비"] += overhead
        standard_total = direct + overhead
        model["standard_total_cost_won"] = standard_total
        model["standard_direct_cost_won"] = direct
        model["standard_overhead_cost_won"] = overhead
        model["gap_vs_reverse_total_won"] = standard_total - model["reverse_total_cost_won"]
        for category, amount in category_totals.items():
            category_rows.append(
                {
                    "survey_type_id": survey_id,
                    "survey_type": model["survey_type"],
                    "cost_category": category,
                    "amount_won": amount,
                }
            )

    return {
        "unit_models": unit_models,
        "labor_roles": roles,
        "cost_categories": category_rows,
        "category_mapping": build_cost_category_mapping(),
        "assumptions": {
            "unit_scope": "1개 측정지점 1회 조사",
            "selected_survey_types": ["기본 현장 유량측정", "자동유량 관측소 운영", "자동유량 관측소 신규 설치"],
            "hourly_rate_basis": "1인 1일 8시간 환산 가정",
            "overhead_rate": 0.17,
            "equipment_life_years": 7,
            "equipment_annual_uses": 160,
        },
    }


def build_database_records() -> dict[str, list[dict[str, Any]]]:
    return {
        "budget_summary": parse_budget_summary_records(),
        "budget_detail": parse_budget_detail_records(),
        "unit_price_status": parse_unit_price_records(),
        "total_budget_breakdown": parse_total_budget_breakdown_records(),
        "residual_transactions": parse_residual_transactions(),
    }


def build_data_quality(records: dict[str, list[dict[str, Any]]], budget: dict[str, Any], flow_breakdown: dict[str, Any]) -> list[dict[str, Any]]:
    checks: list[dict[str, Any]] = []
    summary = records["budget_summary"]
    for year in [year[:4] for year in YEARS]:
        total = next(row for row in summary if row["year"] == year and row["item"] == "계")
        parts = [row for row in summary if row["year"] == year and row["item"] not in {"계"}]
        execution_sum = sum(row["execution_budget_won"] for row in parts)
        carryover_sum = sum(row["balance_reflected_budget_won"] for row in parts)
        checks.append(
            {
                "check": f"{year} 항목별예산 실행예산 합계",
                "expected": total["execution_budget_won"],
                "actual": execution_sum,
                "delta": execution_sum - total["execution_budget_won"],
                "status": "pass" if abs(execution_sum - total["execution_budget_won"]) <= 1 else "review",
            }
        )
        checks.append(
            {
                "check": f"{year} 항목별예산 잔액반영 합계",
                "expected": total["balance_reflected_budget_won"],
                "actual": carryover_sum,
                "delta": carryover_sum - total["balance_reflected_budget_won"],
                "status": "pass" if abs(carryover_sum - total["balance_reflected_budget_won"]) <= 1 else "review",
            }
        )

    flow_budget_2026 = next(row for row in budget["business_budget"] if row["item"] == "유량")["values"]["2026"] * 1_000_000
    checks.append(
        {
            "check": "2026 유량 사업예산 세부 시트 대조",
            "expected": flow_budget_2026,
            "actual": flow_breakdown["total_won"],
            "delta": flow_breakdown["total_won"] - flow_budget_2026,
            "status": "pass" if abs(flow_breakdown["total_won"] - flow_budget_2026) <= 1_000_000 else "review",
        }
    )
    for row in budget["total_budget"]:
        expected = row["total_million"]
        actual = row["org_million"] + row["business_million"]
        checks.append(
            {
                "check": f"{row['year']} 총예산 기관운영+사업",
                "expected": expected,
                "actual": actual,
                "delta": actual - expected,
                "status": "pass" if abs(actual - expected) <= 1 else "review",
            }
        )
    return checks


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_sqlite(records: dict[str, list[dict[str, Any]]], survey_model: dict[str, Any], quality: list[dict[str, Any]]) -> Path:
    db_path = DATA_DIR / "hydrology_budget.sqlite"
    if db_path.exists():
        db_path.unlink()
    conn = sqlite3.connect(db_path)
    try:
        conn.execute(
            """
            CREATE TABLE budget_summary (
                item TEXT, item_with_carryover TEXT, year TEXT,
                execution_budget_won INTEGER, balance_reflected_budget_won INTEGER, balance_delta_won INTEGER,
                source TEXT, source_sheet TEXT, source_row INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO budget_summary VALUES (:item, :item_with_carryover, :year, :execution_budget_won, :balance_reflected_budget_won, :balance_delta_won, :source, :source_sheet, :source_row)",
            records["budget_summary"],
        )
        conn.execute(
            """
            CREATE TABLE budget_detail (
                measurement_item TEXT, budget_group TEXT, stat_item TEXT, year TEXT,
                execution_budget_won INTEGER, carryover_balance_budget_won INTEGER, budget_with_balance_won INTEGER,
                source TEXT, source_sheet TEXT, source_row INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO budget_detail VALUES (:measurement_item, :budget_group, :stat_item, :year, :execution_budget_won, :carryover_balance_budget_won, :budget_with_balance_won, :source, :source_sheet, :source_row)",
            records["budget_detail"],
        )
        conn.execute(
            """
            CREATE TABLE unit_price_status (
                version TEXT, group_name TEXT, item TEXT, year TEXT, count INTEGER,
                unit_price_million_won INTEGER, total_million_won INTEGER, source TEXT, source_sheet TEXT, source_row INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO unit_price_status VALUES (:version, :group, :item, :year, :count, :unit_price_million_won, :total_million_won, :source, :source_sheet, :source_row)",
            records["unit_price_status"],
        )
        conn.execute(
            """
            CREATE TABLE total_budget_breakdown (
                year TEXT, row_category TEXT, item TEXT, amount_million_won INTEGER,
                carryover_included INTEGER, source TEXT, source_sheet TEXT, source_row INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO total_budget_breakdown VALUES (:year, :row_category, :item, :amount_million_won, :carryover_included, :source, :source_sheet, :source_row)",
            [{**row, "carryover_included": int(row["carryover_included"])} for row in records["total_budget_breakdown"]],
        )
        conn.execute(
            """
            CREATE TABLE residual_transactions (
                project_name TEXT, approval_date TEXT, resolution_title TEXT, budget_department TEXT,
                resolution_department TEXT, budget_group TEXT, stat_item TEXT, amount_won INTEGER,
                classification_item TEXT, source TEXT, source_sheet TEXT, source_row INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO residual_transactions VALUES (:project_name, :approval_date, :resolution_title, :budget_department, :resolution_department, :budget_group, :stat_item, :amount_won, :classification_item, :source, :source_sheet, :source_row)",
            records["residual_transactions"],
        )
        conn.execute(
            """
            CREATE TABLE survey_unit_model (
                survey_type_id TEXT, survey_type TEXT, unit_scope TEXT, annual_frequency INTEGER,
                site_count_basis INTEGER, reverse_total_cost_won INTEGER, reverse_business_cost_won INTEGER,
                standard_total_cost_won INTEGER, standard_direct_cost_won INTEGER, standard_overhead_cost_won INTEGER,
                difficulty_level TEXT
            )
            """
        )
        conn.executemany(
            """
            INSERT INTO survey_unit_model VALUES (
                :survey_type_id, :survey_type, :unit_scope, :annual_frequency,
                :site_count_basis, :reverse_total_cost_won, :reverse_business_cost_won,
                :standard_total_cost_won, :standard_direct_cost_won, :standard_overhead_cost_won,
                :difficulty_level
            )
            """,
            survey_model["unit_models"],
        )
        conn.execute(
            """
            CREATE TABLE survey_labor_roles (
                survey_type_id TEXT, survey_type TEXT, role_name TEXT, headcount REAL,
                hours_per_person REAL, labor_grade TEXT, hourly_rate_won INTEGER, labor_cost_won INTEGER,
                cost_category TEXT, basis_method TEXT
            )
            """
        )
        conn.executemany(
            "INSERT INTO survey_labor_roles VALUES (:survey_type_id, :survey_type, :role_name, :headcount, :hours_per_person, :labor_grade, :hourly_rate_won, :labor_cost_won, :cost_category, :basis_method)",
            survey_model["labor_roles"],
        )
        conn.execute(
            """
            CREATE TABLE survey_cost_categories (
                survey_type_id TEXT, survey_type TEXT, cost_category TEXT, amount_won INTEGER
            )
            """
        )
        conn.executemany(
            "INSERT INTO survey_cost_categories VALUES (:survey_type_id, :survey_type, :cost_category, :amount_won)",
            survey_model["cost_categories"],
        )
        conn.execute(
            """
            CREATE TABLE cost_category_mapping (
                standard_cost_category TEXT, budget_group TEXT, stat_item TEXT,
                survey_type TEXT, allocation_rule TEXT
            )
            """
        )
        conn.executemany(
            "INSERT INTO cost_category_mapping VALUES (:standard_cost_category, :budget_group, :stat_item, :survey_type, :allocation_rule)",
            survey_model["category_mapping"],
        )
        conn.execute(
            """
            CREATE TABLE data_quality_checks (
                check_name TEXT, expected REAL, actual REAL, delta REAL, status TEXT
            )
            """
        )
        conn.executemany(
            "INSERT INTO data_quality_checks VALUES (:check, :expected, :actual, :delta, :status)",
            quality,
        )
        conn.commit()
    finally:
        conn.close()
    return db_path


def write_data_exports(data: dict[str, Any], records: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    files = {
        "sqlite": "hydrology_budget.sqlite",
        "summary": "hydrology_budget_summary.json",
        "detail": "hydrology_budget_detail.json",
        "unit_price": "hydrology_budget_unit_price.json",
        "total": "hydrology_budget_total.json",
        "residuals": "hydrology_budget_residuals.json",
        "quality": "hydrology_budget_quality.json",
        "survey_model": "hydrology_survey_unit_model.json",
    }
    write_json(DATA_DIR / files["summary"], records["budget_summary"])
    write_json(DATA_DIR / files["detail"], records["budget_detail"])
    write_json(DATA_DIR / files["unit_price"], records["unit_price_status"])
    write_json(DATA_DIR / files["total"], records["total_budget_breakdown"])
    write_json(DATA_DIR / files["residuals"], records["residual_transactions"])
    write_json(DATA_DIR / files["quality"], data["data_quality"])
    write_json(DATA_DIR / files["survey_model"], data["survey_cost_model"])
    write_sqlite(records, data["survey_cost_model"], data["data_quality"])
    return {
        key: f"data/{filename}"
        for key, filename in files.items()
    }


def enrich_regions(station_data: dict[str, Any], rent_data: dict[str, Any]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = defaultdict(lambda: {"stations": 0, "fares": []})
    for station in station_data["stations"]:
        region = station["region"]
        grouped[region]["stations"] += 1
        if station.get("fare"):
            grouped[region]["fares"].append(station["fare"])

    rows: list[dict[str, Any]] = []
    for region, _, _ in REGION_RANGES:
        fares = grouped[region]["fares"]
        facility = rent_data["container_annual_won"].get(region, 0) + rent_data["parking_annual_won"].get(region, 0)
        count = grouped[region]["stations"]
        rows.append(
            {
                "region": region,
                "stations": count,
                "avg_fare": round(mean(fares)) if fares else 0,
                "min_fare": min(fares) if fares else 0,
                "max_fare": max(fares) if fares else 0,
                "facility_annual_won": facility,
                "facility_per_station_won": round(facility / count) if count else 0,
                "container_annual_won": rent_data["container_annual_won"].get(region, 0),
                "parking_annual_won": rent_data["parking_annual_won"].get(region, 0),
            }
        )
    return rows


def build_data() -> dict[str, Any]:
    budget = parse_budget_and_units()
    flow_breakdown = parse_flow_business_breakdown()
    station_data = parse_stations_and_traffic()
    equipment = parse_equipment()
    vehicles = parse_vehicle_ops()
    rent = parse_rent_ops()
    regions = enrich_regions(station_data, rent)
    db_records = build_database_records()
    survey_cost_model = build_survey_cost_model(budget, flow_breakdown, equipment, vehicles, regions)
    data_quality = build_data_quality(db_records, budget, flow_breakdown)

    flow_2026 = next(row for row in budget["flow_rows"] if row["year"] == "2026")
    total_2026 = next(row for row in budget["total_budget"] if row["year"] == "2026")
    fares = [station["fare"] for station in station_data["stations"] if station.get("fare")]
    pdf_sources = sorted(
        str(path.relative_to(ROOT)).replace("\\", "/") for path in SOURCE_DIR.rglob("*.pdf")
    )
    workbooks = sorted(
        str(path.relative_to(ROOT)).replace("\\", "/")
        for path in SOURCE_DIR.rglob("*.xlsx")
        if not path.name.startswith("~$")
    )
    return {
        "generated_on": date.today().isoformat(),
        "budget": budget,
        "flow_breakdown": flow_breakdown,
        "stations": station_data["stations"],
        "traffic_ranges": station_data["traffic_ranges"],
        "regions": regions,
        "equipment": equipment,
        "vehicles": vehicles,
        "rent": rent,
        "survey_cost_model": survey_cost_model,
        "data_quality": data_quality,
        "database_records": db_records,
        "database_stats": {key: len(value) for key, value in db_records.items()},
        "sources": {
            "workbooks": workbooks,
            "pdfs": pdf_sources,
        },
        "summary": {
            "station_total": len(station_data["stations"]),
            "flow_sites_2026": flow_2026["sites"],
            "flow_unit_price_won": flow_2026["unit_price_million"] * 1_000_000,
            "flow_total_budget_won": flow_2026["total_million"] * 1_000_000,
            "flow_business_budget_won": flow_breakdown["total_won"],
            "flow_business_per_site_won": round(flow_breakdown["total_won"] / flow_2026["sites"]),
            "total_budget_2026_won": total_2026["total_million"] * 1_000_000,
            "total_business_2026_won": total_2026["business_million"] * 1_000_000,
            "avg_one_way_fare_won": round(mean(fares)),
            "max_one_way_fare_won": max(fares),
            "equipment_total": equipment["total_owned"],
            "vehicle_count": vehicles["vehicle_count"],
            "avg_vehicle_rent_won": vehicles["avg_monthly_rent_won"],
            "core_equipment_kit_won": equipment["core_kit_price_won"],
            "workbook_count": len(workbooks),
            "pdf_count": len(pdf_sources),
        },
    }


HTML_TEMPLATE = r"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>수문조사 원가 분석 대시보드</title>
  <style>
    :root {
      --bg: #f5f7fb;
      --panel: #ffffff;
      --ink: #17202a;
      --muted: #667085;
      --line: #d8dee9;
      --blue: #2563eb;
      --teal: #0f766e;
      --green: #15803d;
      --amber: #b45309;
      --violet: #7c3aed;
      --rose: #be123c;
      --soft-blue: #e8f0ff;
      --soft-teal: #e7f7f4;
      --soft-amber: #fff4df;
      --soft-green: #e9f8ee;
      --shadow: 0 10px 28px rgba(15, 23, 42, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      background: var(--bg);
      color: var(--ink);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Noto Sans KR", Arial, sans-serif;
      line-height: 1.5;
    }
    button, input, select {
      font: inherit;
    }
    .app {
      min-height: 100vh;
    }
    .topbar {
      position: sticky;
      top: 0;
      z-index: 10;
      background: rgba(245, 247, 251, 0.94);
      border-bottom: 1px solid var(--line);
      backdrop-filter: blur(12px);
    }
    .topbar-inner {
      max-width: 1480px;
      margin: 0 auto;
      padding: 16px 24px;
      display: grid;
      grid-template-columns: minmax(260px, 1fr) auto;
      gap: 16px;
      align-items: center;
    }
    h1 {
      margin: 0;
      font-size: 22px;
      letter-spacing: 0;
    }
    .subtitle {
      margin-top: 3px;
      color: var(--muted);
      font-size: 13px;
    }
    .actions {
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      justify-content: flex-end;
    }
    .button {
      border: 1px solid var(--line);
      background: var(--panel);
      color: var(--ink);
      padding: 9px 12px;
      border-radius: 8px;
      cursor: pointer;
      min-height: 38px;
      box-shadow: 0 1px 0 rgba(15, 23, 42, 0.03);
    }
    .button.primary {
      background: var(--blue);
      border-color: var(--blue);
      color: white;
    }
    .button:hover {
      border-color: #9aa4b2;
    }
    main {
      max-width: 1480px;
      margin: 0 auto;
      padding: 20px 24px 48px;
    }
    .tabs {
      display: flex;
      gap: 6px;
      overflow-x: auto;
      padding-bottom: 10px;
    }
    .tab {
      border: 1px solid var(--line);
      background: white;
      color: #344054;
      border-radius: 8px;
      padding: 9px 13px;
      white-space: nowrap;
      cursor: pointer;
    }
    .tab.active {
      background: #17202a;
      color: white;
      border-color: #17202a;
    }
    .panel {
      display: none;
    }
    .panel.active {
      display: block;
    }
    .grid {
      display: grid;
      grid-template-columns: repeat(12, minmax(0, 1fr));
      gap: 14px;
      align-items: start;
    }
    .card {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      padding: 16px;
      min-width: 0;
    }
    .span-2 { grid-column: span 2; }
    .span-3 { grid-column: span 3; }
    .span-4 { grid-column: span 4; }
    .span-5 { grid-column: span 5; }
    .span-6 { grid-column: span 6; }
    .span-7 { grid-column: span 7; }
    .span-8 { grid-column: span 8; }
    .span-12 { grid-column: span 12; }
    h2 {
      margin: 0 0 12px;
      font-size: 16px;
      letter-spacing: 0;
    }
    h3 {
      margin: 0 0 8px;
      font-size: 14px;
    }
    .metric-label {
      color: var(--muted);
      font-size: 12px;
    }
    .metric-value {
      margin-top: 6px;
      font-size: 24px;
      font-weight: 800;
      letter-spacing: 0;
      word-break: keep-all;
    }
    .metric-note {
      margin-top: 5px;
      color: var(--muted);
      font-size: 12px;
    }
    .accent-blue { border-top: 4px solid var(--blue); }
    .accent-teal { border-top: 4px solid var(--teal); }
    .accent-green { border-top: 4px solid var(--green); }
    .accent-amber { border-top: 4px solid var(--amber); }
    .accent-violet { border-top: 4px solid var(--violet); }
    .chart {
      width: 100%;
      height: 310px;
      display: block;
      background: #fbfcff;
      border: 1px solid #e5eaf2;
      border-radius: 8px;
    }
    .chart.small {
      height: 250px;
    }
    .chart-canvas {
      width: 100%;
      height: 310px;
      display: block;
      background: #fbfcff;
      border: 1px solid #e5eaf2;
      border-radius: 8px;
      padding: 12px;
      position: relative;
    }
    .chart-canvas.small {
      height: 250px;
    }
    .chart-canvas canvas {
      width: 100% !important;
      height: 100% !important;
    }
    .table-chart-grid {
      display: grid;
      grid-template-columns: minmax(300px, 1.05fr) minmax(260px, 0.95fr);
      gap: 12px;
      align-items: stretch;
    }
    .legend {
      display: flex;
      flex-wrap: wrap;
      gap: 10px 14px;
      margin-top: 10px;
      color: var(--muted);
      font-size: 12px;
    }
    .legend i {
      display: inline-block;
      width: 10px;
      height: 10px;
      border-radius: 3px;
      margin-right: 5px;
      vertical-align: -1px;
    }
    .table-wrap {
      overflow: auto;
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 13px;
      min-width: 680px;
    }
    th, td {
      padding: 10px;
      border-bottom: 1px solid #edf0f5;
      text-align: left;
      vertical-align: top;
    }
    th {
      color: #344054;
      background: #f8fafc;
      font-weight: 700;
    }
    tr:last-child td {
      border-bottom: 0;
    }
    .form-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
    }
    label {
      display: block;
      color: var(--muted);
      font-size: 12px;
      margin-bottom: 5px;
    }
    input, select {
      width: 100%;
      border: 1px solid #cfd7e3;
      border-radius: 8px;
      padding: 9px 10px;
      background: white;
      color: var(--ink);
      min-height: 38px;
    }
    .summary-line {
      display: flex;
      justify-content: space-between;
      gap: 10px;
      padding: 10px 0;
      border-bottom: 1px solid #edf0f5;
      font-size: 13px;
    }
    .summary-line:last-child {
      border-bottom: 0;
    }
    .amount {
      font-weight: 800;
      text-align: right;
      white-space: nowrap;
    }
    .status-band {
      display: grid;
      grid-template-columns: 1.2fr 1fr 1fr;
      gap: 12px;
      align-items: stretch;
    }
    .callout {
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 13px;
      background: #fbfcff;
    }
    .callout strong {
      display: block;
      margin-bottom: 4px;
    }
    .soft-blue { background: var(--soft-blue); }
    .soft-teal { background: var(--soft-teal); }
    .soft-green { background: var(--soft-green); }
    .soft-amber { background: var(--soft-amber); }
    .pill {
      display: inline-block;
      border: 1px solid #cfd7e3;
      border-radius: 999px;
      padding: 3px 8px;
      margin: 2px;
      font-size: 12px;
      color: #344054;
      background: white;
      white-space: nowrap;
    }
    .station-strip {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 10px;
      margin-top: 12px;
    }
    .mini-map {
      min-height: 310px;
      display: grid;
      place-items: center;
      background: linear-gradient(180deg, #fbfcff, #f2f7ff);
      border: 1px solid #e5eaf2;
      border-radius: 8px;
    }
    .source-list {
      columns: 2;
      column-gap: 24px;
      padding-left: 18px;
      margin: 0;
      color: #344054;
      font-size: 13px;
    }
    .source-list li {
      break-inside: avoid;
      margin-bottom: 8px;
    }
    .export-grid {
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
    }
    .export-link {
      display: block;
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 10px;
      color: var(--ink);
      text-decoration: none;
      background: #fbfcff;
      font-size: 13px;
    }
    .export-link strong {
      display: block;
      margin-bottom: 4px;
    }
    .print-title {
      display: none;
    }
    @media (max-width: 1100px) {
      .span-2, .span-3, .span-4, .span-5, .span-6, .span-7, .span-8 { grid-column: span 12; }
      .form-grid, .station-strip, .status-band, .export-grid, .table-chart-grid { grid-template-columns: 1fr 1fr; }
      .topbar-inner { grid-template-columns: 1fr; }
      .actions { justify-content: flex-start; }
    }
    @media (max-width: 640px) {
      main { padding: 14px 12px 34px; }
      .topbar-inner { padding: 14px 12px; }
      .form-grid, .station-strip, .status-band, .export-grid, .table-chart-grid { grid-template-columns: 1fr; }
      .source-list { columns: 1; }
      .metric-value { font-size: 21px; }
      h1 { font-size: 19px; }
    }
    @media print {
      body { background: white; }
      .topbar, .tabs, .no-print { display: none !important; }
      main { max-width: none; padding: 0; }
      .panel { display: block !important; page-break-after: always; }
      .card { box-shadow: none; break-inside: avoid; }
      .print-title { display: block; margin: 0 0 16px; }
    }
  </style>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.9/dist/chart.umd.min.js"></script>
</head>
<body>
  <script type="application/json" id="dashboard-data">__DATA_JSON__</script>
  <div class="app">
    <header class="topbar">
      <div class="topbar-inner">
        <div>
          <h1>수문조사 원가 분석 대시보드</h1>
          <div class="subtitle" id="subtitle"></div>
        </div>
        <div class="actions no-print">
          <button class="button" id="btnCsv">CSV</button>
          <button class="button" id="btnExcel">Excel</button>
          <button class="button primary" id="btnPrint">인쇄/PDF</button>
        </div>
      </div>
    </header>
    <main>
      <div class="tabs no-print" id="tabs"></div>

      <section class="panel active" id="overview">
        <h2 class="print-title">개요</h2>
        <div class="grid" id="overviewGrid"></div>
      </section>

      <section class="panel" id="unit-cost">
        <h2 class="print-title">1회 조사 원가모델</h2>
        <div class="grid">
          <div class="card span-12">
            <h2>분석 조건</h2>
            <div class="form-grid">
              <div>
                <label for="surveyTypeSelect">조사 유형</label>
                <select id="surveyTypeSelect"></select>
                <div class="metric-note">사용자 선택: 기본 현장 유량측정, 자동유량 운영, 자동유량 신규 설치</div>
              </div>
              <div>
                <label for="surveyFrequencyInput">연간 조사 횟수</label>
                <input id="surveyFrequencyInput" type="number" min="1" step="1" value="12">
                <div class="metric-note">예산 역산 원가는 횟수 변경에 따라 1회 단가로 재배분</div>
              </div>
              <div>
                <label for="difficultyInput">현장 난이도 배율</label>
                <input id="difficultyInput" type="number" min="0.5" step="0.1" value="1">
                <div class="metric-note">역할별 투입시간에 적용</div>
              </div>
              <div>
                <label for="modelOverheadInput">간접비율</label>
                <input id="modelOverheadInput" type="number" min="0" step="1" value="17">
                <div class="metric-note">표준 직접비 기준</div>
              </div>
            </div>
          </div>
          <div class="card span-3 accent-blue" id="unitCostPeople"></div>
          <div class="card span-3 accent-teal" id="unitCostHours"></div>
          <div class="card span-3 accent-green" id="unitCostStandard"></div>
          <div class="card span-3 accent-amber" id="unitCostReverse"></div>
          <div class="card span-6">
            <h2>표준 모델 vs 예산 역산</h2>
            <div class="chart-canvas"><canvas id="unitCostCompareChart"></canvas></div>
          </div>
          <div class="card span-6">
            <h2>역할별 투입 인시</h2>
            <div class="chart-canvas"><canvas id="laborHoursChart"></canvas></div>
          </div>
          <div class="card span-6">
            <h2>비용 구성</h2>
            <div class="chart-canvas small"><canvas id="costStructureChart"></canvas></div>
          </div>
          <div class="card span-6">
            <h2>연도별 1개소 1회 평균비용</h2>
            <div class="form-grid" style="margin-bottom:10px">
              <div>
                <label for="trendViewSelect">차트 보기</label>
                <select id="trendViewSelect">
                  <option value="average">조사유형 평균</option>
                  <option value="station">지점별 비교</option>
                </select>
              </div>
              <div>
                <label for="trendStationSelect">비교 지점</label>
                <select id="trendStationSelect"></select>
              </div>
            </div>
            <div class="chart-canvas small"><canvas id="perSiteTrendChart"></canvas></div>
            <div class="metric-note" id="perSiteTrendNote" style="margin-top:8px"></div>
          </div>
          <div class="card span-12">
            <h2>역할별 산정표</h2>
            <div class="table-wrap"><table id="unitCostRoleTable"></table></div>
          </div>
        </div>
      </section>

      <section class="panel" id="station">
        <h2 class="print-title">지점별 원가 조회</h2>
        <div class="grid">
          <div class="card span-12">
            <h2>지점 선택 및 산출 조건</h2>
            <div class="form-grid">
              <div>
                <label for="stationSelect">조사지점</label>
                <select id="stationSelect"></select>
                <div class="metric-note">출처: 조사지점 현황_v1.xlsx</div>
              </div>
              <div>
                <label for="staffInput">투입 인원</label>
                <input id="staffInput" type="number" min="1" value="2" title="기준: 수문조사 2인 1조 표준. 지점 난이도에 따라 조정 가능">
                <div class="metric-note">기준: 2인 1조 표준 (지점 조건별 조정)</div>
              </div>
              <div>
                <label for="visitInput">연간 방문 횟수</label>
                <input id="visitInput" type="number" min="1" value="12" title="기준: 월 1회 = 연 12회 방문. 실측 빈도에 따라 조정">
                <div class="metric-note">기준: 월 1회 연 12회 (빈도별 조정)</div>
              </div>
              <div>
                <label for="dayInput">방문당 투입일</label>
                <input id="dayInput" type="number" min="0.25" step="0.25" value="1" title="기준: 지점당 1일 투입 (당일 조사)">
                <div class="metric-note">기준: 1일 (원거리 지점 1.5~2일)</div>
              </div>
              <div>
                <label for="laborInput">1인 1일 인건비</label>
                <input id="laborInput" type="number" min="0" step="10000" value="250000" title="기준: 건설업 보통인부 노임단가 참고 (고용노동부·대한건설협회 발표값). 직종별 차이 있음">
                <div class="metric-note">기준: 노임단가표 보통인부 기준 (직종별 차이 있음)</div>
              </div>
              <div>
                <label for="vehicleSlotInput">차량 월 처리 지점</label>
                <input id="vehicleSlotInput" type="number" min="1" value="22" title="기준: 차량 1대가 월평균 처리하는 지점 수 추정값. 실제 배치 현황(업무차량 workbook)과 지점 수 비교 필요">
                <div class="metric-note">기준: 추정값 — 업무차량 workbook 배치 현황 참고</div>
              </div>
              <div>
                <label for="equipmentLifeInput">장비 내용연수</label>
                <input id="equipmentLifeInput" type="number" min="1" value="7" title="기준: 유량측정 장비 평균 내용연수 7년 (장비 관련 workbook 정수기준 참고)">
                <div class="metric-note">기준: 장비 관련 workbook · 보유장비 관련 시트 정수기준</div>
              </div>
              <div>
                <label for="equipmentUseInput">장비 세트 연간 투입 횟수</label>
                <input id="equipmentUseInput" type="number" min="1" value="160" title="기준: 장비 세트 연간 조사 투입 횟수 추정 (144개소 × 월 1회 기준)">
                <div class="metric-note">기준: 지점 수 × 방문 횟수 기준 추정</div>
              </div>
              <div>
                <label for="overheadInput">간접비율</label>
                <input id="overheadInput" type="number" min="0" step="1" value="17" title="기준: 수자원조사 표준 간접비율 17% (세부예산 현황 참고)">
                <div class="metric-note">기준: 표준 간접비율 17% (세부예산 현황 참고)</div>
              </div>
            </div>
          </div>
          <div class="card span-4 accent-blue" id="stationInfo"></div>
          <div class="card span-4 accent-teal" id="stationCost"></div>
          <div class="card span-4 accent-amber" id="stationCompare"></div>
          <div class="card span-7">
            <h2>지점 원가 구성</h2>
            <svg class="chart small" id="costDonut" viewBox="0 0 720 250" role="img"></svg>
            <div class="legend" id="costLegend"></div>
          </div>
          <div class="card span-5">
            <h2>산출 내역</h2>
            <div id="costLines"></div>
          </div>
          <div class="card span-12">
            <h2>교통비 상위 지점</h2>
            <div class="table-wrap"><table id="topFareTable"></table></div>
            <div class="metric-note" style="margin-top:8px">출처: 교통비 workbook (조사지점 현황_v1.xlsx 연번 매칭)</div>
          </div>
        </div>
      </section>

      <section class="panel" id="region">
        <h2 class="print-title">권역별 비교</h2>
        <div class="grid">
          <div class="card span-12">
            <h2>권역 요약</h2>
            <div class="station-strip" id="regionCards"></div>
            <div class="metric-note" style="margin-top:8px">출처: 조사지점 현황_v1.xlsx · 교통비 workbook · 임대 관련 workbook (컨테이너·주차 시트)</div>
          </div>
          <div class="card span-6">
            <h2>평균 편도 교통비</h2>
            <svg class="chart small" id="regionFareChart" viewBox="0 0 720 250" role="img"></svg>
            <div class="metric-note" style="margin-top:8px">출처: 교통비 workbook · 권역별 평균값 (최소~최대 상세는 비교표 참조)</div>
          </div>
          <div class="card span-6">
            <h2>연간 임차·주차비 지점 배부액</h2>
            <svg class="chart small" id="regionFacilityChart" viewBox="0 0 720 250" role="img"></svg>
            <div class="metric-note" style="margin-top:8px">출처: 임대 관련 workbook · 컨테이너(창고)·주차시설 시트 (2025년 기준, 연간합계 ÷ 지점수)</div>
          </div>
          <div class="card span-12">
            <h2>권역별 비교표</h2>
            <div class="table-wrap"><table id="regionTable"></table></div>
            <div class="metric-note" style="margin-top:8px">출처: 교통비 workbook (편도교통비) · 임대 관련 workbook (컨테이너·주차) · 2025년 기준</div>
          </div>
        </div>
      </section>

      <section class="panel" id="equipment">
        <h2 class="print-title">장비·운영비</h2>
        <div class="grid">
          <div class="card span-3 accent-blue" id="equipmentKpi"></div>
          <div class="card span-3 accent-teal" id="vehicleKpi"></div>
          <div class="card span-3 accent-green" id="calibrationKpi"></div>
          <div class="card span-3 accent-amber" id="rentKpi"></div>
          <div class="card span-7">
            <h2>보유 장비 현황</h2>
            <div class="table-wrap"><table id="equipmentTable"></table></div>
            <div class="metric-note" style="margin-top:8px">출처: 장비 관련 workbook · 보유장비 관련 시트 (행5~, 열B~J) · 정수기준·단가 포함</div>
          </div>
          <div class="card span-5">
            <h2>차량·에너지 운영</h2>
            <div id="vehicleOps"></div>
            <div class="metric-note" style="margin-top:8px">출처: 업무차량 workbook · 업무차량 시트 / 전기&amp;주유비 시트</div>
          </div>
          <div class="card span-12">
            <h2>월별 차량 에너지 운영비</h2>
            <svg class="chart small" id="energyChart" viewBox="0 0 820 250" role="img"></svg>
            <div class="legend" id="energyLegend"></div>
            <div class="metric-note" style="margin-top:8px">출처: 업무차량 workbook · 전기&amp;주유비 시트 (행4~5, 열C~H)</div>
          </div>
        </div>
      </section>

      <section class="panel" id="sources">
        <h2 class="print-title">자료 출처</h2>
        <div class="grid">
          <div class="card span-6">
            <h2>Excel 원자료</h2>
            <ul class="source-list" id="workbookSources"></ul>
          </div>
          <div class="card span-6">
            <h2>지침·법령 PDF</h2>
            <ul class="source-list" id="pdfSources"></ul>
          </div>
          <div class="card span-12">
            <h2>데이터 처리 기준 및 수치 근거</h2>
            <div class="status-band">
              <div class="callout soft-blue"><strong>권역 분류</strong>조사지점 연번 흐름에 따라 1~97 한강, 98~201 낙동강, 202~278 금강, 279~355 영산강으로 배부했습니다. (출처: 조사지점 현황_v1.xlsx, 행4~)</div>
              <div class="callout soft-teal"><strong>교통비 매칭</strong>교통비 파일의 연번 범위를 조사지점 번호로 펼쳐 355개 지점 모두에 편도 교통비를 연결했습니다. (출처: 교통비 workbook, 열A~E)</div>
              <div class="callout soft-amber"><strong>원가 계산</strong>지점별 계산기는 실제 교통비, 권역별 임차비, 차량 평균 임차료, 대표 장비 세트 감가상각을 조합한 추정 모델입니다. 각 항목의 기본값 근거는 지점별 탭 폼 필드 하단에 명시되어 있습니다.</div>
            </div>
            <div style="height:12px"></div>
            <div class="status-band">
              <div class="callout soft-green"><strong>단가 현황</strong>단가 및 수량 데이터는 수문조사 항목별 예산 및 단가 현황_v2.xlsx의 단가 시트(행5~11)와 현황 시트(행5~11)에서 추출. 단가 보기/수량 보기 토글로 전환 가능.</div>
              <div class="callout soft-blue"><strong>유량 세부예산</strong>세부예산 현황_v1.xlsx · 유량 시트에서 2026년 소계 항목(인건비·운영비·여비·업무추진비·유형자산) 합산. 각 항목 근거는 유량 사업 세부예산 분해 카드에 명시.</div>
              <div class="callout soft-teal"><strong>보유 장비</strong>장비 관련 workbook · 보유장비 관련 시트(행5~, 열B~J). 단가는 실구매가 또는 견적가 기준. 지역별 배치 현황은 해당 workbook 내 별도 시트 확인 필요.</div>
            </div>
          </div>
          <div class="card span-12">
            <h2>DB 산출물 및 검증</h2>
            <div class="export-grid" id="dataExportList"></div>
            <div style="height:12px"></div>
            <div class="table-wrap"><table id="dataQualityTable"></table></div>
          </div>
        </div>
      </section>
    </main>
  </div>

  <script>
    const DATA = JSON.parse(document.getElementById("dashboard-data").textContent);
    const TABS = [
      ["overview", "개요"],
      ["unit-cost", "1회 조사"],
      ["station", "지점별"],
      ["region", "권역 비교"],
      ["equipment", "장비·운영비"],
      ["sources", "자료 출처"]
    ];
    const COLORS = ["#2563eb", "#0f766e", "#15803d", "#b45309", "#7c3aed", "#be123c"];
    const CHARTS = {};
    const $ = (id) => document.getElementById(id);
    const nf = new Intl.NumberFormat("ko-KR");
    const won = (v) => `${nf.format(Math.round(Number(v) || 0))}원`;
    const billion = (v) => `${(Number(v || 0) / 100000000).toLocaleString("ko-KR", { maximumFractionDigits: 1 })}억원`;
    const million = (v) => `${nf.format(Math.round(Number(v || 0) / 1000000))}백만원`;
    const pct = (v) => `${Number(v || 0).toFixed(1)}%`;
    const escapeHtml = (value) => String(value ?? "").replace(/[&<>"']/g, (ch) => ({ "&": "&amp;", "<": "&lt;", ">": "&gt;", '"': "&quot;", "'": "&#39;" }[ch]));

    function initTabs() {
      $("tabs").innerHTML = TABS.map(([id, label], idx) => `<button class="tab ${idx === 0 ? "active" : ""}" data-tab="${id}">${label}</button>`).join("");
      $("tabs").addEventListener("click", (event) => {
        const button = event.target.closest("[data-tab]");
        if (!button) return;
        document.querySelectorAll(".tab").forEach((tab) => tab.classList.toggle("active", tab === button));
        document.querySelectorAll(".panel").forEach((panel) => panel.classList.toggle("active", panel.id === button.dataset.tab));
        if (button.dataset.tab === "unit-cost") {
          setTimeout(renderUnitCostModel, 0);
        }
      });
    }

    function metric(label, value, note, accent) {
      return `<div class="card span-3 ${accent || ""}"><div class="metric-label">${label}</div><div class="metric-value">${value}</div><div class="metric-note">${note}</div></div>`;
    }

    function renderOverview() {
      const s = DATA.summary;
      $("subtitle").textContent = `원자료 ${s.workbook_count}개 workbook, ${s.pdf_count}개 PDF 기준 · 생성일 ${DATA.generated_on}`;
      const latestFlow = DATA.budget.flow_rows.find((row) => row.year === "2026");
      $("overviewGrid").innerHTML = `
        ${metric("2026 유량 지점 수", `${nf.format(s.flow_sites_2026)}개소`, "유량조사 기준", "accent-blue")}
        ${metric("유량 지점당 단가", billion(s.flow_unit_price_won), "2021~2026 고정", "accent-teal")}
        ${metric("2026 수문정보 총예산", billion(s.total_budget_2026_won), "총계 기준", "accent-green")}
        ${metric("평균 편도 교통비", won(s.avg_one_way_fare_won), "355개 지점 매칭", "accent-amber")}
        <div class="card span-8">
          <h2>수문정보 총예산 추이</h2>
          <svg class="chart" id="totalBudgetChart" viewBox="0 0 900 310" role="img"></svg>
          <div class="legend"><span><i style="background:#2563eb"></i>기관운영</span><span><i style="background:#0f766e"></i>사업</span></div>
        </div>
        <div class="card span-4">
          <h2>2026 유량 예산 구조</h2>
          <div class="summary-line"><span>총계</span><span class="amount">${million(latestFlow.total_million * 1000000)}</span></div>
          <div class="summary-line"><span>기관운영</span><span class="amount">${million(latestFlow.org_million * 1000000)} · ${pct(latestFlow.org_pct)}</span></div>
          <div class="summary-line"><span>사업</span><span class="amount">${million(latestFlow.business_million * 1000000)} · ${pct(latestFlow.business_pct)}</span></div>
          <div class="summary-line"><span>사업 예산 지점 배부</span><span class="amount">${won(s.flow_business_per_site_won)}</span></div>
        </div>
        <div class="card span-5">
          <h2>항목별 사업 예산</h2>
          <svg class="chart small" id="itemBudgetChart" viewBox="0 0 820 250" role="img"></svg>
        </div>
        <div class="card span-7">
          <h2>단가 및 수량 현황</h2>
          <div style="margin-bottom:8px;display:flex;gap:6px">
            <button class="button" id="unitViewToggle" data-view="price" style="font-size:12px;padding:4px 10px;min-height:28px">수량 보기</button>
          </div>
          <div class="table-chart-grid">
            <div class="table-wrap"><table id="unitPriceTable"></table></div>
            <div class="chart-canvas small"><canvas id="unitStatusChart"></canvas></div>
          </div>
          <div class="metric-note" style="margin-top:8px">출처: 수문조사 항목별 예산 및 단가 현황_v2.xlsx · 단가 시트 (행5~11), 현황 시트 (행5~11)</div>
        </div>
        <div class="card span-6">
          <h2>유량 사업 세부예산 분해</h2>
          <div id="flowBreakdown"></div>
          <div class="chart-canvas small" style="margin-top:12px"><canvas id="flowBreakdownChart"></canvas></div>
        </div>
        <div class="card span-6">
          <h2>수문조사 지점·자동유량 수량</h2>
          <div class="chart-canvas small"><canvas id="countChart"></canvas></div>
        </div>
      `;
      stackedBudgetChart("totalBudgetChart", DATA.budget.total_budget);
      barChart("itemBudgetChart", DATA.budget.business_budget.filter((row) => row.item !== "계").map((row) => ({ label: row.item, value: row.values["2026"] * 1000000 })), "budget");
      renderUnitTable();
      renderFlowBreakdown();
      renderCountQuantityChart();
    }

    function stackedBudgetChart(id, rows) {
      const width = 900, height = 310, left = 70, right = 24, top = 28, bottom = 44;
      const innerW = width - left - right, innerH = height - top - bottom;
      const max = Math.max(...rows.map((row) => row.total_million));
      const band = innerW / rows.length;
      let html = `<line x1="${left}" y1="${height - bottom}" x2="${width - right}" y2="${height - bottom}" stroke="#98a2b3"/>`;
      rows.forEach((row, i) => {
        const x = left + i * band + band * 0.22;
        const bw = band * 0.56;
        const orgH = row.org_million / max * innerH;
        const bizH = row.business_million / max * innerH;
        const base = height - bottom;
        html += `<rect x="${x}" y="${base - orgH}" width="${bw}" height="${orgH}" rx="4" fill="#2563eb"/>`;
        html += `<rect x="${x}" y="${base - orgH - bizH}" width="${bw}" height="${bizH}" rx="4" fill="#0f766e"/>`;
        html += `<text x="${x + bw / 2}" y="${base + 22}" text-anchor="middle" font-size="13" fill="#667085">${row.year}</text>`;
        html += `<text x="${x + bw / 2}" y="${base - orgH - bizH - 8}" text-anchor="middle" font-size="12" font-weight="700" fill="#17202a">${nf.format(row.total_million)}</text>`;
      });
      $("totalBudgetChart").innerHTML = html;
    }

    function barChart(id, rows, kind) {
      const width = 820, height = 250, left = 120, right = 28, top = 18, bottom = 28;
      const innerW = width - left - right, innerH = height - top - bottom;
      const MIN_LABEL_SPACE = 80; // pixels needed after bar end to show label outside
      const max = Math.max(...rows.map((row) => row.value), 1);
      const band = innerH / rows.length;
      let html = `<line x1="${left}" y1="${height - bottom}" x2="${width - right}" y2="${height - bottom}" stroke="#d0d5dd"/>`;
      rows.forEach((row, i) => {
        const y = top + i * band + band * 0.2;
        const bh = band * 0.6;
        const bw = row.value / max * innerW;
        html += `<text x="${left - 10}" y="${y + bh * 0.68}" text-anchor="end" font-size="13" fill="#344054">${escapeHtml(row.label)}</text>`;
        html += `<rect x="${left}" y="${y}" width="${bw}" height="${bh}" rx="5" fill="${COLORS[i % COLORS.length]}"/>`;
        const label = kind === "budget" ? billion(row.value) : kind === "won" ? won(row.value) : nf.format(row.value);
        const remainingSpace = innerW - bw;
        if (remainingSpace < MIN_LABEL_SPACE) {
          html += `<text x="${left + bw - 6}" y="${y + bh * 0.68}" text-anchor="end" font-size="12" font-weight="700" fill="white">${label}</text>`;
        } else {
          html += `<text x="${left + bw + 8}" y="${y + bh * 0.68}" font-size="12" font-weight="700" fill="#17202a">${label}</text>`;
        }
      });
      $(id).innerHTML = html;
    }

    function groupedLineChart(id, rows) {
      const width = 820, height = 250, left = 48, right = 24, top = 24, bottom = 36;
      const years = ["2021", "2022", "2023", "2024", "2025", "2026"];
      const max = Math.max(...rows.flatMap((row) => years.map((year) => row.values[year] || 0)), 1);
      const x = (idx) => left + idx * ((width - left - right) / (years.length - 1));
      const y = (value) => height - bottom - (value / max) * (height - top - bottom);
      let html = `<line x1="${left}" y1="${height - bottom}" x2="${width - right}" y2="${height - bottom}" stroke="#d0d5dd"/>`;
      years.forEach((year, idx) => {
        html += `<text x="${x(idx)}" y="${height - 12}" text-anchor="middle" font-size="12" fill="#667085">${year}</text>`;
      });
      rows.forEach((row, seriesIdx) => {
        const points = years.map((year, idx) => `${x(idx)},${y(row.values[year] || 0)}`).join(" ");
        html += `<polyline fill="none" stroke="${COLORS[seriesIdx]}" stroke-width="3" points="${points}"/>`;
        years.forEach((year, idx) => {
          html += `<circle cx="${x(idx)}" cy="${y(row.values[year] || 0)}" r="4" fill="${COLORS[seriesIdx]}"/>`;
        });
        html += `<text x="${left + seriesIdx * 90}" y="${top - 8}" font-size="12" fill="${COLORS[seriesIdx]}" font-weight="700">${escapeHtml(row.item)}</text>`;
      });
      $(id).innerHTML = html;
    }

    function renderUnitTable(view) {
      const years = ["2021", "2022", "2023", "2024", "2025", "2026"];
      const btn = $("unitViewToggle");
      if (!view) {
        view = (btn && btn.dataset.view) ? btn.dataset.view : "price";
      }
      let rows, headerSuffix;
      if (view === "count") {
        rows = DATA.budget.counts;
        headerSuffix = " (수량·개소)";
        if (btn) { btn.textContent = "단가 보기"; btn.dataset.view = "count"; }
      } else {
        rows = DATA.budget.unit_prices;
        headerSuffix = " (단가·백만원/개소)";
        if (btn) { btn.textContent = "수량 보기"; btn.dataset.view = "price"; }
      }
      const tbody = rows.map((row) => {
        const countRow = DATA.budget.counts.find((c) => c.item === row.item);
        const cells = years.map((year) => {
          if (view === "count") {
            return `<td>${nf.format(row.values[year])}개소</td>`;
          }
          const qty = countRow ? countRow.values[year] : null;
          const tipText = qty != null
            ? `${row.values[year]}백만원 × ${qty}개소 = ${nf.format(row.values[year] * qty)}백만원`
            : "";
          const tip = tipText ? ` title="${tipText}"` : "";
          const qtyNote = qty != null ? `<br><span class='metric-note'>${qty}개소</span>` : "";
          return `<td${tip}>${nf.format(row.values[year])}백만원${qtyNote}</td>`;
        }).join("");
        return `<tr><td>${escapeHtml(row.item)}</td>${cells}</tr>`;
      }).join("");
      $("unitPriceTable").innerHTML = `<thead><tr><th>항목${escapeHtml(headerSuffix)}</th>${years.map((year) => `<th>${year}</th>`).join("")}</tr></thead><tbody>${tbody}</tbody>`;
      renderUnitStatusChart(view);
    }

    function initUnitViewToggle() {
      const btn = $("unitViewToggle");
      if (!btn) return;
      btn.addEventListener("click", () => {
        const next = btn.dataset.view === "price" ? "count" : "price";
        btn.dataset.view = next;
        renderUnitTable(next);
      });
    }

    function renderFlowBreakdown() {
      const rows = DATA.flow_breakdown.rows;
      const total = DATA.flow_breakdown.total_won;
      const ITEM_BASIS = {
        "인건비": "관리직·기능직 노임단가 기준 — 출처: 세부예산 현황_v1.xlsx · 유량 시트",
        "운영비": "조사 운영 소모품 등 직접비 — 출처: 세부예산 현황_v1.xlsx · 유량 시트",
        "여비": "355개 지점 교통비 합산 — 출처: 세부예산 현황_v1.xlsx · 유량 시트",
        "업무추진비": "내부 기준 비율 적용 — 출처: 세부예산 현황_v1.xlsx · 유량 시트",
        "유형자산": "장비 취득·교체 계획 — 출처: 세부예산 현황_v1.xlsx · 유량 시트",
      };
      $("flowBreakdown").innerHTML = rows.map((row, idx) => `
        <div class="summary-line">
          <span><i style="display:inline-block;width:10px;height:10px;border-radius:3px;background:${COLORS[idx % COLORS.length]};margin-right:6px"></i>${escapeHtml(row.item)}</span>
          <span class="amount">${billion(row.amount_won)} · ${pct(row.amount_won / total * 100)}</span>
        </div>
        ${ITEM_BASIS[row.item] ? `<div class="metric-note" style="padding:0 0 6px 16px">${escapeHtml(ITEM_BASIS[row.item])}</div>` : ""}
      `).join("") + `<div class="metric-note" style="margin-top:6px;border-top:1px solid #edf0f5;padding-top:6px">2026 유량 사업 예산 ${billion(total)} 기준 · 출처: ${escapeHtml(DATA.flow_breakdown.source)} · 유량 시트</div>`;
      renderFlowBreakdownChart();
    }

    function makeChart(id, config) {
      const canvas = $(id);
      if (!canvas) return;
      if (!window.Chart) {
        const holder = canvas.closest(".chart-canvas");
        if (holder) {
          holder.innerHTML = '<div class="metric-note">Chart.js를 불러오지 못했습니다. 표와 DB 산출물을 확인하세요.</div>';
        }
        return;
      }
      if (CHARTS[id]) {
        CHARTS[id].destroy();
      }
      CHARTS[id] = new Chart(canvas, config);
    }

    function chartInteractionOptions() {
      return {
        mode: "nearest",
        intersect: true,
      };
    }

    function moneyTooltip(context) {
      let value = context.parsed;
      if (value && typeof value === "object") value = value.y ?? value.x;
      return `${context.dataset.label || context.label}: ${won(value)}`;
    }

    function renderUnitStatusChart(view) {
      const years = ["2021", "2022", "2023", "2024", "2025", "2026"];
      const isCount = view === "count";
      const rows = isCount ? DATA.budget.counts : DATA.budget.unit_prices;
      makeChart("unitStatusChart", {
        type: "line",
        data: {
          labels: years,
          datasets: rows.map((row, idx) => ({
            label: row.item,
            data: years.map((year) => row.values[year] || 0),
            borderColor: COLORS[idx % COLORS.length],
            backgroundColor: COLORS[idx % COLORS.length],
            pointRadius: 3,
            pointHoverRadius: 7,
            tension: 0.25,
          })),
        },
        options: {
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteractionOptions(),
          plugins: {
            legend: { position: "bottom" },
            tooltip: {
              callbacks: {
                label: (ctx) => {
                  const row = rows[ctx.datasetIndex];
                  const year = years[ctx.dataIndex];
                  const value = ctx.parsed.y || 0;
                  if (isCount) {
                    return `${row.item}: ${nf.format(value)}개소`;
                  }
                  const countRow = DATA.budget.counts.find((item) => item.item === row.item);
                  const count = countRow ? countRow.values[year] : null;
                  const total = count == null ? null : value * count;
                  return [
                    `${row.item}: ${nf.format(value)}백만원/개소`,
                    count == null ? "" : `수량 ${nf.format(count)}개소 · 산출액 ${nf.format(total)}백만원`,
                  ].filter(Boolean);
                },
              },
            },
          },
          scales: {
            y: {
              ticks: {
                callback: (value) => isCount ? `${nf.format(value)}개소` : `${nf.format(value)}백만원`,
              },
            },
          },
        },
      });
    }

    function renderFlowBreakdownChart() {
      const rows = DATA.flow_breakdown.rows;
      const total = rows.reduce((sum, row) => sum + Number(row.amount_won || 0), 0) || 1;
      makeChart("flowBreakdownChart", {
        type: "doughnut",
        data: {
          labels: rows.map((row) => row.item),
          datasets: [{
            label: "세부예산",
            data: rows.map((row) => row.amount_won),
            backgroundColor: COLORS,
            borderColor: "#ffffff",
            borderWidth: 2,
          }],
        },
        options: {
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteractionOptions(),
          plugins: {
            legend: { position: "bottom" },
            tooltip: {
              callbacks: {
                label: (ctx) => {
                  const value = Number(ctx.parsed || 0);
                  return `${ctx.label}: ${billion(value)} · ${pct(value / total * 100)}`;
                },
              },
            },
          },
        },
      });
    }

    function renderCountQuantityChart() {
      const years = ["2021", "2022", "2023", "2024", "2025", "2026"];
      const rows = DATA.budget.counts.filter((row) => ["유량", "운영", "설치"].includes(row.item));
      makeChart("countChart", {
        type: "line",
        data: {
          labels: years,
          datasets: rows.map((row, idx) => ({
            label: row.item,
            data: years.map((year) => row.values[year] || 0),
            borderColor: COLORS[idx % COLORS.length],
            backgroundColor: COLORS[idx % COLORS.length],
            pointRadius: 4,
            pointHoverRadius: 8,
            tension: 0.25,
          })),
        },
        options: {
          responsive: true,
          maintainAspectRatio: false,
          interaction: chartInteractionOptions(),
          plugins: {
            legend: { position: "bottom" },
            tooltip: {
              callbacks: {
                label: (ctx) => `${ctx.dataset.label}: ${nf.format(ctx.parsed.y || 0)}개소`,
              },
            },
          },
          scales: {
            y: {
              ticks: { callback: (value) => `${nf.format(value)}개소` },
            },
          },
        },
      });
    }


    function initSurveyModelControls() {
      const select = $("surveyTypeSelect");
      if (!select || !DATA.survey_cost_model) return;
      select.innerHTML = DATA.survey_cost_model.unit_models.map((model) => `<option value="${model.survey_type_id}">${escapeHtml(model.survey_type)}</option>`).join("");
      const basic = DATA.survey_cost_model.unit_models.find((model) => model.survey_type_id === "basic_flow") || DATA.survey_cost_model.unit_models[0];
      select.value = basic.survey_type_id;
      $("surveyFrequencyInput").value = basic.annual_frequency || 1;
      if ($("trendStationSelect")) {
        $("trendStationSelect").innerHTML = DATA.stations.map((station) => `<option value="${station.no}">${station.no}. ${escapeHtml(station.name)} · ${escapeHtml(station.region)} · ${won(station.fare)}</option>`).join("");
      }
      if ($("trendViewSelect")) {
        $("trendViewSelect").addEventListener("change", () => {
          updateTrendStationControl();
          renderUnitCostModel();
        });
      }
      if ($("trendStationSelect")) {
        $("trendStationSelect").addEventListener("change", renderUnitCostModel);
      }
      select.addEventListener("change", () => {
        const model = selectedSurveyModel();
        $("surveyFrequencyInput").value = model.annual_frequency || 1;
        renderUnitCostModel();
      });
      ["surveyFrequencyInput", "difficultyInput", "modelOverheadInput"].forEach((id) => {
        $(id).addEventListener("input", renderUnitCostModel);
        $(id).addEventListener("change", renderUnitCostModel);
      });
      updateTrendStationControl();
    }

    function selectedSurveyModel() {
      const models = DATA.survey_cost_model?.unit_models || [];
      const id = $("surveyTypeSelect")?.value || models[0]?.survey_type_id;
      return models.find((model) => model.survey_type_id === id) || models[0];
    }

    function computeSurveyCostModel() {
      const model = selectedSurveyModel();
      if (!model) return null;
      const frequency = Math.max(1, Number($("surveyFrequencyInput").value || model.annual_frequency || 1));
      const difficulty = Math.max(0.1, Number($("difficultyInput").value || 1));
      const overheadRate = Math.max(0, Number($("modelOverheadInput").value || 0)) / 100;
      const roles = DATA.survey_cost_model.labor_roles.filter((row) => row.survey_type_id === model.survey_type_id);
      const roleCategories = new Set(roles.map((row) => row.cost_category));
      const adjustedRoles = roles.map((row) => {
        const adjustedHours = Number(row.headcount || 0) * Number(row.hours_per_person || 0) * difficulty;
        return {
          ...row,
          adjusted_hours: adjustedHours,
          adjusted_cost_won: Math.round(adjustedHours * Number(row.hourly_rate_won || 0)),
        };
      });
      const categoryTotals = {};
      adjustedRoles.forEach((row) => {
        categoryTotals[row.cost_category] = (categoryTotals[row.cost_category] || 0) + row.adjusted_cost_won;
      });
      const sourceCategories = DATA.survey_cost_model.cost_categories.filter((row) => row.survey_type_id === model.survey_type_id);
      sourceCategories.forEach((row) => {
        const name = String(row.cost_category || "");
        if (roleCategories.has(row.cost_category) || name.includes("간접")) return;
        categoryTotals[row.cost_category] = (categoryTotals[row.cost_category] || 0) + Number(row.amount_won || 0);
      });
      const direct = Object.values(categoryTotals).reduce((sum, value) => sum + Number(value || 0), 0);
      const overhead = Math.round(direct * overheadRate);
      categoryTotals["간접비"] = overhead;
      const total = Math.round(direct + overhead);
      const baseFrequency = Number(model.annual_frequency || frequency || 1);
      const reverseTotal = Math.round(Number(model.reverse_total_cost_won || 0) * baseFrequency / frequency);
      const reverseBusiness = Math.round(Number(model.reverse_business_cost_won || 0) * baseFrequency / frequency);
      return {
        model,
        frequency,
        difficulty,
        overheadRate,
        adjustedRoles,
        categoryRows: Object.entries(categoryTotals).map(([cost_category, amount_won]) => ({ cost_category, amount_won })),
        people: adjustedRoles.reduce((sum, row) => sum + Number(row.headcount || 0), 0),
        hours: adjustedRoles.reduce((sum, row) => sum + Number(row.adjusted_hours || 0), 0),
        direct,
        overhead,
        total,
        reverseTotal,
        reverseBusiness,
        gap: total - reverseTotal,
      };
    }

    function updateTrendStationControl() {
      const stationSelect = $("trendStationSelect");
      if (!stationSelect) return;
      const stationMode = $("trendViewSelect")?.value === "station";
      stationSelect.disabled = !stationMode;
      stationSelect.style.opacity = stationMode ? "1" : "0.55";
    }

    function selectedTrendStation() {
      const no = Number($("trendStationSelect")?.value || DATA.stations[0]?.no);
      return DATA.stations.find((station) => station.no === no) || DATA.stations[0];
    }

    function surveyTrendRows(model, frequencyOverride) {
      const years = ["2021", "2022", "2023", "2024", "2025", "2026"];
      if (!model) return [];
      const overrideFrequency = Math.max(1, Number(frequencyOverride || model.annual_frequency || 1));
      if (model.survey_type_id === "basic_flow") {
        return years.map((year) => {
          const row = DATA.budget.flow_rows.find((item) => item.year === year);
          return { year, value: row ? Math.round(row.total_million * 1000000 / row.sites / overrideFrequency) : 0 };
        });
      }
      const keyword = model.survey_type_id === "auto_install" ? "설치" : "운영";
      const frequency = model.survey_type_id === "auto_install" ? 1 : overrideFrequency;
      return years.map((year) => {
        const row = DATA.budget.auto_rows.find((item) => item.year === year && String(item.item).includes(keyword));
        return { year, value: row ? Math.round(row.total_million * 1000000 / Math.max(row.quantity, 1) / frequency) : 0 };
      });
    }

    function stationTrendBasis(model, frequency) {
      const station = selectedTrendStation();
      const region = DATA.regions.find((row) => row.region === station.region) || DATA.regions[0];
      const stationCount = DATA.regions.reduce((sum, row) => sum + Number(row.stations || 0), 0) || DATA.regions.length || 1;
      const avgFacility = DATA.regions.reduce((sum, row) => sum + Number(row.facility_per_station_won || 0) * Number(row.stations || 1), 0) / stationCount;
      const averageRoundTrip = Math.max(Number(DATA.summary.avg_one_way_fare_won || 0) * 2, 1);
      const modelTravel = Number(model.travel_cost_won || 0);
      const travelMultiplier = modelTravel > 0 ? modelTravel / averageRoundTrip : 1;
      const stationTravel = Number(station.fare || 0) * 2 * travelMultiplier;
      const facilityDeltaPerVisit = (Number(region.facility_per_station_won || 0) - avgFacility) / Math.max(1, Number(frequency || model.annual_frequency || 1));
      const adjustment = stationTravel - modelTravel + facilityDeltaPerVisit;
      return {
        station,
        region,
        stationTravel,
        modelTravel,
        facilityDeltaPerVisit,
        adjustment,
      };
    }

    function stationTrendRows(model, frequency, averageRows) {
      const basis = stationTrendBasis(model, frequency);
      return {
        basis,
        rows: averageRows.map((row) => ({
          year: row.year,
          value: Math.max(0, Math.round(row.value + basis.adjustment)),
        })),
      };
    }

    function renderUnitCostModel() {
      if (!DATA.survey_cost_model || !$("surveyTypeSelect")) return;
      const calc = computeSurveyCostModel();
      if (!calc) return;
      $("unitCostPeople").innerHTML = `<div class="metric-label">역할 투입 인원</div><div class="metric-value">${nf.format(calc.people)}명</div><div class="metric-note">역할별 투입 합산</div>`;
      $("unitCostHours").innerHTML = `<div class="metric-label">총 투입 인시</div><div class="metric-value">${calc.hours.toLocaleString("ko-KR", { maximumFractionDigits: 1 })}h</div><div class="metric-note">난이도 ${calc.difficulty.toFixed(1)}배 적용</div>`;
      $("unitCostStandard").innerHTML = `<div class="metric-label">표준 모델 원가</div><div class="metric-value">${won(calc.total)}</div><div class="metric-note">직접비 ${won(calc.direct)} + 간접비 ${won(calc.overhead)}</div>`;
      $("unitCostReverse").innerHTML = `<div class="metric-label">예산 역산 평균</div><div class="metric-value">${won(calc.reverseTotal)}</div><div class="metric-note">총액 기준 · 사업비 기준 ${won(calc.reverseBusiness)}</div>`;
      $("unitCostRoleTable").innerHTML = `<thead><tr><th>역할</th><th>인원</th><th>1인 시간</th><th>적용 인시</th><th>시간당 단가</th><th>역할 비용</th><th>범주</th></tr></thead><tbody>${calc.adjustedRoles.map((row) => `<tr><td>${escapeHtml(row.role_name)}</td><td>${nf.format(row.headcount)}명</td><td>${Number(row.hours_per_person).toFixed(1)}h</td><td>${row.adjusted_hours.toLocaleString("ko-KR", { maximumFractionDigits: 1 })}h</td><td>${won(row.hourly_rate_won)}</td><td>${won(row.adjusted_cost_won)}</td><td>${escapeHtml(row.cost_category)}</td></tr>`).join("")}</tbody>`;
      renderSurveyCharts(calc);
      renderDataExports();
    }

    function renderSurveyCharts(calc) {
      const commonOptions = {
        responsive: true,
        maintainAspectRatio: false,
        plugins: { legend: { position: "bottom" }, tooltip: { callbacks: { label: moneyTooltip } } },
      };
      makeChart("unitCostCompareChart", {
        type: "bar",
        data: {
          labels: ["표준 모델", "예산 역산(총액)", "예산 역산(사업)"],
          datasets: [{ label: "1회 원가", data: [calc.total, calc.reverseTotal, calc.reverseBusiness], backgroundColor: ["#2563eb", "#0f766e", "#b45309"] }],
        },
        options: { ...commonOptions, scales: { y: { ticks: { callback: (value) => won(value) } } } },
      });
      makeChart("laborHoursChart", {
        type: "bar",
        data: {
          labels: calc.adjustedRoles.map((row) => row.role_name),
          datasets: [{ label: "투입 인시", data: calc.adjustedRoles.map((row) => row.adjusted_hours), backgroundColor: COLORS }],
        },
        options: {
          responsive: true,
          maintainAspectRatio: false,
          indexAxis: "y",
          plugins: { legend: { display: false }, tooltip: { callbacks: { label: (ctx) => `${ctx.parsed.x.toLocaleString("ko-KR", { maximumFractionDigits: 1 })}h` } } },
          scales: { x: { ticks: { callback: (value) => `${value}h` } } },
        },
      });
      makeChart("costStructureChart", {
        type: "doughnut",
        data: {
          labels: calc.categoryRows.map((row) => row.cost_category),
          datasets: [{ label: "비용", data: calc.categoryRows.map((row) => row.amount_won), backgroundColor: COLORS }],
        },
        options: commonOptions,
      });
      const trendFrequency = calc.model.survey_type_id === "auto_install" ? 1 : calc.frequency;
      const trend = surveyTrendRows(calc.model, trendFrequency);
      const stationMode = $("trendViewSelect")?.value === "station";
      const stationTrend = stationMode ? stationTrendRows(calc.model, trendFrequency, trend) : null;
      const trendDatasets = [
        { label: "조사유형 평균", data: trend.map((row) => row.value), borderColor: "#2563eb", backgroundColor: "rgba(37, 99, 235, 0.12)", tension: 0.25, fill: true },
      ];
      if (stationTrend) {
        trendDatasets.push({ label: `${stationTrend.basis.station.name} 추정`, data: stationTrend.rows.map((row) => row.value), borderColor: "#b45309", backgroundColor: "rgba(180, 83, 9, 0.08)", tension: 0.25, fill: false });
      }
      if ($("perSiteTrendNote")) {
        $("perSiteTrendNote").textContent = stationTrend
          ? `${stationTrend.basis.station.name} · ${stationTrend.basis.region.region} · 편도 교통비 ${won(stationTrend.basis.station.fare)} 기준으로 평균 추이에 교통비/권역 시설 배부 차이를 반영했습니다.`
          : "조사유형별 전체 예산을 개소 수와 연간 조사 횟수로 나눈 1개소 1회 평균비용입니다.";
      }
      makeChart("perSiteTrendChart", {
        type: "line",
        data: {
          labels: trend.map((row) => row.year),
          datasets: trendDatasets,
        },
        options: {
          ...commonOptions,
          interaction: { mode: "nearest", intersect: true },
          plugins: {
            legend: { position: "bottom" },
            tooltip: {
              callbacks: {
                label: moneyTooltip,
                afterBody: () => stationTrend ? [
                  `지점: ${stationTrend.basis.station.name}`,
                  `편도 교통비: ${won(stationTrend.basis.station.fare)}`,
                  `1회 보정: ${won(stationTrend.basis.adjustment)}`,
                ] : [],
              },
            },
          },
          scales: { y: { ticks: { callback: (value) => won(value) } } },
        },
      });
    }

    function renderDataExports() {
      if (!$("dataExportList")) return;
      const labels = {
        sqlite: "SQLite DB",
        summary: "항목별 예산 JSON",
        detail: "세부예산 JSON",
        unit_price: "단가·수량 JSON",
        total: "총예산 JSON",
        residuals: "집행잔액 JSON",
        quality: "검증 JSON",
        survey_model: "1회 조사 모델 JSON",
      };
      const statKeys = {
        summary: "budget_summary",
        detail: "budget_detail",
        unit_price: "unit_price_status",
        total: "total_budget_breakdown",
        residuals: "residual_transactions",
      };
      $("dataExportList").innerHTML = Object.entries(DATA.data_exports || {}).map(([key, path]) => {
        const count = DATA.database_stats?.[statKeys[key] || key];
        const countLabel = count == null ? "" : `${nf.format(count)} rows`;
        return `<a class="export-link" href="${escapeHtml(path)}"><strong>${escapeHtml(labels[key] || key)}</strong><span>${escapeHtml(path)}</span><br><span class="metric-note">${countLabel}</span></a>`;
      }).join("");
      const checks = DATA.data_quality || [];
      $("dataQualityTable").innerHTML = `<thead><tr><th>검증 항목</th><th>기대값</th><th>실제값</th><th>차이</th><th>상태</th></tr></thead><tbody>${checks.map((row) => `<tr><td>${escapeHtml(row.check)}</td><td>${nf.format(Math.round(row.expected || 0))}</td><td>${nf.format(Math.round(row.actual || 0))}</td><td>${nf.format(Math.round(row.delta || 0))}</td><td><span class="pill">${escapeHtml(row.status)}</span></td></tr>`).join("")}</tbody>`;
    }

    function initStationSelect() {
      $("stationSelect").innerHTML = DATA.stations.map((station) => `<option value="${station.no}">${station.no}. ${escapeHtml(station.name)} · ${escapeHtml(station.region)} · ${won(station.fare)}</option>`).join("");
      ["stationSelect", "staffInput", "visitInput", "dayInput", "laborInput", "vehicleSlotInput", "equipmentLifeInput", "equipmentUseInput", "overheadInput"].forEach((id) => {
        $(id).addEventListener("input", renderStation);
        $(id).addEventListener("change", renderStation);
      });
    }

    function selectedStation() {
      const no = Number($("stationSelect").value || DATA.stations[0].no);
      return DATA.stations.find((station) => station.no === no) || DATA.stations[0];
    }

    function calcStationCost() {
      const station = selectedStation();
      const region = DATA.regions.find((row) => row.region === station.region);
      const staff = Number($("staffInput").value || 0);
      const visits = Number($("visitInput").value || 0);
      const days = Number($("dayInput").value || 0);
      const laborDaily = Number($("laborInput").value || 0);
      const vehicleSlots = Number($("vehicleSlotInput").value || 1);
      const equipmentLife = Number($("equipmentLifeInput").value || 1);
      const equipmentUses = Number($("equipmentUseInput").value || 1);
      const overheadRate = Number($("overheadInput").value || 0);
      const labor = staff * days * laborDaily * visits;
      const traffic = station.fare * 2 * staff * visits;
      const vehicle = DATA.summary.avg_vehicle_rent_won / vehicleSlots * visits;
      const facility = region.facility_per_station_won;
      const equipment = DATA.summary.core_equipment_kit_won / equipmentLife / equipmentUses * visits;
      const direct = labor + traffic + vehicle + facility + equipment;
      const overhead = direct * overheadRate / 100;
      const total = direct + overhead;
      return {
        station, region, staff, visits, days, laborDaily, vehicleSlots, equipmentLife, equipmentUses, overheadRate,
        lines: [
          ["직접인건비", labor, `${staff}명 × ${days}일 × ${won(laborDaily)} × ${visits}회`],
          ["교통비", traffic, `${won(station.fare)} × 왕복 × ${staff}명 × ${visits}회`],
          ["차량 운영비", vehicle, `${won(DATA.summary.avg_vehicle_rent_won)} ÷ ${vehicleSlots}지점/월 × ${visits}회`],
          ["창고·주차 임차 배부", facility, `${station.region} 연간 임차·주차비 ÷ ${region.stations}개 지점`],
          ["장비 감가상각", equipment, `${won(DATA.summary.core_equipment_kit_won)} ÷ ${equipmentLife}년 ÷ ${equipmentUses}회 × ${visits}회`],
          ["간접비", overhead, `직접비 × ${overheadRate}%`],
        ],
        direct, overhead, total,
        businessPerSite: DATA.summary.flow_business_per_site_won,
        unitPrice: DATA.summary.flow_unit_price_won
      };
    }

    function renderStation() {
      const calc = calcStationCost();
      $("stationInfo").innerHTML = `
        <div class="metric-label">선택 지점</div>
        <div class="metric-value">${escapeHtml(calc.station.name)}</div>
        <div class="metric-note">${escapeHtml(calc.station.address)}</div>
        <div class="summary-line"><span>권역</span><span class="amount">${escapeHtml(calc.station.region)}</span></div>
        <div class="summary-line"><span>거점</span><span class="amount">${escapeHtml(calc.station.hub)}</span></div>
        <div class="summary-line"><span>교통수단</span><span class="amount">${escapeHtml(calc.station.mode)}</span></div>
      `;
      $("stationCost").innerHTML = `
        <div class="metric-label">연간 지점 추정 원가</div>
        <div class="metric-value">${billion(calc.total)}</div>
        <div class="metric-note">선택 조건 기준</div>
        <div class="summary-line"><span>직접비</span><span class="amount">${won(calc.direct)}</span></div>
        <div class="summary-line"><span>간접비</span><span class="amount">${won(calc.overhead)}</span></div>
      `;
      const businessRatio = calc.total / calc.businessPerSite * 100;
      const unitRatio = calc.total / calc.unitPrice * 100;
      $("stationCompare").innerHTML = `
        <div class="metric-label">예산 기준 대비</div>
        <div class="metric-value">${pct(businessRatio)}</div>
        <div class="metric-note">유량 사업예산 지점 배부액 대비</div>
        <div class="summary-line"><span>사업 배부 기준</span><span class="amount">${won(calc.businessPerSite)}</span></div>
        <div class="summary-line"><span>84백만원 단가 대비</span><span class="amount">${pct(unitRatio)}</span></div>
      `;
      donutChart(calc.lines.map((line) => ({ label: line[0], value: line[1] })));
      $("costLines").innerHTML = calc.lines.map(([label, value, formula]) => `
        <div class="summary-line"><span>${escapeHtml(label)}<br><span class="metric-note">${escapeHtml(formula)}</span></span><span class="amount">${won(value)}</span></div>
      `).join("");
    }

    function donutChart(rows) {
      const total = rows.reduce((sum, row) => sum + row.value, 0) || 1;
      const cx = 170, cy = 125, r = 78, circumference = 2 * Math.PI * r;
      let offset = 0;
      let svg = `<circle cx="${cx}" cy="${cy}" r="${r}" fill="none" stroke="#edf0f5" stroke-width="34"/>`;
      rows.forEach((row, idx) => {
        const dash = row.value / total * circumference;
        svg += `<circle cx="${cx}" cy="${cy}" r="${r}" fill="none" stroke="${COLORS[idx % COLORS.length]}" stroke-width="34" stroke-dasharray="${dash} ${circumference - dash}" stroke-dashoffset="${-offset}" transform="rotate(-90 ${cx} ${cy})"/>`;
        offset += dash;
      });
      svg += `<text x="${cx}" y="${cy - 4}" text-anchor="middle" font-size="16" font-weight="800" fill="#17202a">${billion(total)}</text><text x="${cx}" y="${cy + 18}" text-anchor="middle" font-size="12" fill="#667085">연간 추정</text>`;
      rows.forEach((row, idx) => {
        const x = 330, y = 44 + idx * 30;
        svg += `<rect x="${x}" y="${y - 11}" width="12" height="12" rx="3" fill="${COLORS[idx % COLORS.length]}"/><text x="${x + 20}" y="${y}" font-size="13" fill="#344054">${escapeHtml(row.label)}</text><text x="${x + 210}" y="${y}" font-size="13" font-weight="700" fill="#17202a">${pct(row.value / total * 100)}</text>`;
      });
      $("costDonut").innerHTML = svg;
      $("costLegend").innerHTML = rows.map((row, idx) => `<span><i style="background:${COLORS[idx % COLORS.length]}"></i>${escapeHtml(row.label)} ${won(row.value)}</span>`).join("");
    }

    function renderTopFares() {
      const rows = [...DATA.stations].sort((a, b) => b.fare - a.fare).slice(0, 12);
      $("topFareTable").innerHTML = `<thead><tr><th>연번</th><th>관측소</th><th>권역</th><th>편도 교통비</th><th>거점</th><th>주소</th></tr></thead><tbody>${rows.map((row) => `<tr><td>${row.no}</td><td>${escapeHtml(row.name)}</td><td>${escapeHtml(row.region)}</td><td>${won(row.fare)}</td><td>${escapeHtml(row.hub)}</td><td>${escapeHtml(row.address)}</td></tr>`).join("")}</tbody>`;
    }

    function renderRegion() {
      $("regionCards").innerHTML = DATA.regions.map((row, idx) => `
        <div class="callout ${["soft-blue", "soft-teal", "soft-green", "soft-amber"][idx]}">
          <strong>${escapeHtml(row.region)}</strong>
          <div class="summary-line"><span>지점</span><span class="amount">${nf.format(row.stations)}개</span></div>
          <div class="summary-line"><span>평균 편도</span><span class="amount">${won(row.avg_fare)}</span></div>
          <div class="summary-line"><span>시설 배부</span><span class="amount">${won(row.facility_per_station_won)}</span></div>
        </div>
      `).join("");
      barChart("regionFareChart", DATA.regions.map((row) => ({ label: row.region, value: row.avg_fare })), "won");
      barChart("regionFacilityChart", DATA.regions.map((row) => ({ label: row.region, value: row.facility_per_station_won })), "won");
      $("regionTable").innerHTML = `<thead><tr><th>권역</th><th>지점 수</th><th>평균 편도</th><th>최소~최대 편도</th><th>컨테이너 임차</th><th>주차비</th><th>지점 배부</th></tr></thead><tbody>${DATA.regions.map((row) => `<tr><td>${escapeHtml(row.region)}</td><td>${nf.format(row.stations)}개</td><td>${won(row.avg_fare)}</td><td>${won(row.min_fare)} ~ ${won(row.max_fare)}</td><td>${won(row.container_annual_won)}</td><td>${won(row.parking_annual_won)}</td><td>${won(row.facility_per_station_won)}</td></tr>`).join("")}</tbody>`;
    }

    function renderEquipment() {
      $("equipmentKpi").innerHTML = `<div class="metric-label">보유 장비</div><div class="metric-value">${nf.format(DATA.summary.equipment_total)}대</div><div class="metric-note">출처: 장비 관련 workbook · 보유장비 관련 시트</div>`;
      $("vehicleKpi").innerHTML = `<div class="metric-label">업무 차량</div><div class="metric-value">${nf.format(DATA.summary.vehicle_count)}대</div><div class="metric-note">월 평균 임차료 ${won(DATA.summary.avg_vehicle_rent_won)}<br>출처: 업무차량 workbook · 업무차량 시트</div>`;
      $("calibrationKpi").innerHTML = `<div class="metric-label">검·교정 연간 비용</div><div class="metric-value">${billion(DATA.equipment.calibration_annual_won)}</div><div class="metric-note">출처: 장비 관련 workbook · 월 검·교정 비용 시트 (행4, 열C~N)</div>`;
      const rentTotal = Object.values(DATA.rent.container_annual_won).reduce((a, b) => a + b, 0) + Object.values(DATA.rent.parking_annual_won).reduce((a, b) => a + b, 0) + DATA.rent.survey_equipment_rent_won;
      $("rentKpi").innerHTML = `<div class="metric-label">임대·주차·측량 대여</div><div class="metric-value">${billion(rentTotal)}</div><div class="metric-note">2025년 기준 합계<br>출처: 임대 관련 workbook · 컨테이너·주차·측량장비 시트</div>`;
      $("equipmentTable").innerHTML = `<thead><tr><th>장비</th><th>보유</th><th>정수 기준</th><th>2027 구매계획</th><th>단가</th></tr></thead><tbody>${DATA.equipment.rows.map((row) => `<tr><td>${escapeHtml(row.name)}</td><td>${nf.format(row.owned)}대</td><td>${escapeHtml(row.standard)}</td><td>${nf.format(row.purchase_plan_2027)}대</td><td>${row.unit_price_won ? won(row.unit_price_won) : "-"}</td></tr>`).join("")}</tbody>`;
      const energyRows = DATA.vehicles.energy_rows.map((row) => `<h3>${escapeHtml(row.item)}</h3>${Object.entries(row.values).map(([month, value]) => `<div class="summary-line"><span>${escapeHtml(month)}</span><span class="amount">${won(value)}</span></div>`).join("")}`).join("");
      const vehicleTypes = DATA.vehicles.type_counts.slice(0, 8).map((row) => `<span class="pill">${escapeHtml(row.type)} ${row.count}대</span>`).join("");
      $("vehicleOps").innerHTML = `<div class="callout soft-blue"><strong>차종 구성</strong>${vehicleTypes}</div><div style="height:12px"></div>${energyRows}`;
      renderEnergyChart();
    }

    function renderEnergyChart() {
      const energyChartEl = $("energyChart");
      if (!energyChartEl) return;
      const allRows = DATA.vehicles.energy_rows;
      if (!allRows || allRows.length === 0) return;
      const LEGEND_ITEM_SPACING = 130; // pixels between legend entries in the chart header
      const width = 820, height = 250, left = 60, right = 24, top = 28, bottom = 40;
      const innerW = width - left - right, innerH = height - top - bottom;
      const months = Object.keys(allRows[0].values);
      const band = innerW / months.length;
      const barW = band * 0.38;
      const allValues = allRows.reduce((acc, row) => acc.concat(Object.values(row.values)), []);
      const max = allValues.reduce((a, b) => Math.max(a, b), 1);
      const yScale = (v) => height - bottom - (v / max) * innerH;
      let html = `<line x1="${left}" y1="${height - bottom}" x2="${width - right}" y2="${height - bottom}" stroke="#d0d5dd"/>`;
      months.forEach((month, mi) => {
        const cx = left + mi * band + band / 2;
        html += `<text x="${cx}" y="${height - 12}" text-anchor="middle" font-size="12" fill="#667085">${escapeHtml(month)}</text>`;
        allRows.forEach((row, ri) => {
          const value = row.values[month] || 0;
          const x = cx - barW + ri * barW;
          const barH = (value / max) * innerH;
          const y = yScale(value);
          html += `<rect x="${x}" y="${y}" width="${barW - 2}" height="${barH}" rx="3" fill="${COLORS[ri % COLORS.length]}"/>`;
        });
      });
      allRows.forEach((row, ri) => {
        const total = Object.values(row.values).reduce((a, b) => a + b, 0);
        html += `<text x="${left + ri * LEGEND_ITEM_SPACING}" y="${top - 8}" font-size="12" font-weight="700" fill="${COLORS[ri % COLORS.length]}">${escapeHtml(row.item)} (합계 ${won(total)})</text>`;
      });
      energyChartEl.innerHTML = html;
      $("energyLegend").innerHTML = allRows.map((row, ri) => `<span><i style="background:${COLORS[ri % COLORS.length]}"></i>${escapeHtml(row.item)}</span>`).join("");
    }

    function renderSources() {
      $("workbookSources").innerHTML = DATA.sources.workbooks.map((path) => `<li>${escapeHtml(path)}</li>`).join("");
      $("pdfSources").innerHTML = DATA.sources.pdfs.map((path) => `<li>${escapeHtml(path)}</li>`).join("");
      renderDataExports();
    }

    function currentExportRows() {
      const calc = calcStationCost();
      const rows = [
        ["구분", "항목", "값", "근거"],
        ["선택 지점", "연번", calc.station.no, calc.station.name],
        ["선택 지점", "권역", calc.station.region, calc.station.address],
        ["선택 지점", "편도 교통비", Math.round(calc.station.fare), calc.station.hub],
        ...calc.lines.map(([label, value, formula]) => ["원가", label, Math.round(value), formula]),
        ["원가", "연간 지점 추정 원가", Math.round(calc.total), "직접비 + 간접비"],
        ["비교", "유량 사업예산 지점 배부액", Math.round(calc.businessPerSite), "2026 유량 사업예산 / 144개소"],
        ["비교", "유량 지점당 단가", Math.round(calc.unitPrice), "84백만원"]
      ];
      DATA.regions.forEach((row) => rows.push(["권역", row.region, row.avg_fare, `지점 ${row.stations}개, 시설 배부 ${row.facility_per_station_won}원`]));
      const unitCost = computeSurveyCostModel();
      if (unitCost) {
        rows.push(["1회 조사", "조사 유형", unitCost.model.survey_type, unitCost.model.unit_scope]);
        rows.push(["1회 조사", "연간 조사 횟수", unitCost.frequency, "예산 역산 배부 기준"]);
        rows.push(["1회 조사", "총 투입 인시", unitCost.hours.toFixed(1), "역할별 투입시간 합산"]);
        rows.push(["1회 조사", "표준 모델 원가", Math.round(unitCost.total), "역할별 투입 + 비용 범주 + 간접비"]);
        rows.push(["1회 조사", "예산 역산 평균 총액", Math.round(unitCost.reverseTotal), "과거/현재 예산 단가 역산"]);
        unitCost.categoryRows.forEach((row) => rows.push(["1회 조사 비용구성", row.cost_category, Math.round(row.amount_won), "표준 원가모델 범주"]));
      }
      return rows;
    }

    function downloadCsv() {
      const csv = currentExportRows().map((row) => row.map((value) => `"${String(value).replaceAll('"', '""')}"`).join(",")).join("\n");
      const blob = new Blob(["\ufeff" + csv], { type: "text/csv;charset=utf-8" });
      triggerDownload(blob, "수문조사_원가분석_현재뷰.csv");
    }

    function downloadExcel() {
      const rows = currentExportRows();
      const html = `<html><head><meta charset="utf-8"></head><body><table>${rows.map((row, idx) => `<tr>${row.map((value) => {
        const tag = idx === 0 ? "th" : "td";
        return `<${tag}>${escapeHtml(value)}</${tag}>`;
      }).join("")}</tr>`).join("")}</table></body></html>`;
      const blob = new Blob(["\ufeff" + html], { type: "application/vnd.ms-excel;charset=utf-8" });
      triggerDownload(blob, "수문조사_원가분석_현재뷰.xls");
    }

    function triggerDownload(blob, filename) {
      const a = document.createElement("a");
      a.href = URL.createObjectURL(blob);
      a.download = filename;
      a.click();
      URL.revokeObjectURL(a.href);
    }

    function bindActions() {
      $("btnCsv").addEventListener("click", downloadCsv);
      $("btnExcel").addEventListener("click", downloadExcel);
      $("btnPrint").addEventListener("click", () => window.print());
    }

    function init() {
      initTabs();
      renderOverview();
      initUnitViewToggle();
      initSurveyModelControls();
      renderUnitCostModel();
      initStationSelect();
      renderStation();
      renderTopFares();
      renderRegion();
      renderEquipment();
      renderSources();
      bindActions();
    }

    init();
  </script>
</body>
</html>
"""


def main() -> None:
    data = build_data()
    database_records = data.pop("database_records")
    data["data_exports"] = write_data_exports(data, database_records)
    data_json = json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    OUTPUT_HTML.write_text(HTML_TEMPLATE.replace("__DATA_JSON__", data_json), encoding="utf-8")
    print(f"Wrote {OUTPUT_HTML}")
    print(
        json.dumps(
            {
                "stations": data["summary"]["station_total"],
                "flow_sites_2026": data["summary"]["flow_sites_2026"],
                "avg_one_way_fare_won": data["summary"]["avg_one_way_fare_won"],
                "equipment_total": data["summary"]["equipment_total"],
                "vehicles": data["summary"]["vehicle_count"],
                "database_rows": sum(data["database_stats"].values()),
                "data_exports": len(data["data_exports"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
